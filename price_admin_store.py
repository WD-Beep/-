from __future__ import annotations

import base64
import json
import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from kb_data_quality import (
    KB_ACTION_AUTO,
    KB_ACTION_DROP,
    KB_ACTION_REVIEW,
    KbDataQualityVerdict,
    classify_exception_review_hint,
    format_exception_reason_label,
    judge_kb_insert_candidate,
)
from core.knowledge_reload import KNOWLEDGE_MUTATION_LOCK, knowledge_reload_hook
from price_kb import format_material_unit_price_text, get_price_kb, note_kb_disk_write_success
from price_kb_paths import (
    LEGACY_EXCEPTION_PATH,
    admin_price_meta,
    assert_official_kb_write_allowed,
    drop_log_path,
    exception_path as review_exception_file,
    history_path,
    is_official_kb_path,
    official_kb_path,
    auto_learn_log_path,
    quote_sync_suggestions_path,
)
from price_kb_pollution import (
    filter_visible_exceptions,
    is_test_price_exception_record,
    is_test_quote_sync_context,
)
from sheet_parser import normalize_rows, parse_sheet_xml_rows, read_sheet_entries, read_shared_strings


ROOT = Path(__file__).resolve().parent
from price_kb_paths import OFFICIAL_KB_PATH_DEFAULT  # noqa: E402

# 兼容旧 import；列表/读盘请用 official_kb_path()。
DEFAULT_KB_PATH = OFFICIAL_KB_PATH_DEFAULT


def _kb_target(kb_path: Path | None = None) -> Path:
    return (kb_path or official_kb_path()).resolve()


def _exc_target(exc_path: Path | None = None) -> Path:
    return (exc_path or review_exception_file()).resolve()


def _hist_target(hist_path: Path | None = None) -> Path:
    return (hist_path or history_path()).resolve()


def _drop_target(log_path: Path | None = None) -> Path:
    return (log_path or drop_log_path()).resolve()


def _load_merged_exception_entries() -> list[dict[str, Any]]:
    """合并待审核队列文件（新目录 + 遗留 data/），去重 exception_id。"""
    paths: list[Path] = []
    for p in (review_exception_file(), LEGACY_EXCEPTION_PATH):
        rp = p.resolve()
        if rp.is_file() and rp not in paths:
            paths.append(rp)
    merged: dict[str, dict[str, Any]] = {}
    for p in paths:
        for row in _read_exception_entries(p):
            eid = str(row.get("exception_id") or row.get("row_id") or "").strip()
            key = eid or f"__row_{len(merged)}_{row.get('name')}"
            merged[key] = row
    return list(merged.values())


def _count_auto_learn_log_entries() -> int:
    p = auto_learn_log_path()
    if not p.is_file():
        return 0
    try:
        return sum(1 for line in p.read_text(encoding="utf-8").splitlines() if line.strip())
    except OSError:
        return 0


def _append_auto_learn_log(record: dict[str, Any]) -> None:
    path = auto_learn_log_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def _count_quote_sync_suggestions() -> int:
    p = quote_sync_suggestions_path()
    if not p.is_file():
        return 0
    try:
        return sum(1 for line in p.read_text(encoding="utf-8").splitlines() if line.strip())
    except OSError:
        return 0


def _append_quote_sync_suggestions(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    path = quote_sync_suggestions_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    saved: list[dict[str, Any]] = []
    with path.open("a", encoding="utf-8") as fh:
        for raw in rows:
            rec = dict(raw)
            rec.setdefault("queued_at", now)
            rec.setdefault("status", "pending_review")
            rec.setdefault("source", "quote_auto_sync")
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
            saved.append(rec)
    return saved

_KB_SLASH_PRICE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*/\s*[A-Za-z?]+\s*", re.I)
_KB_YUAN_UNIT_PRICE_RE = re.compile(r"元\s*/", re.I)


def _preserve_kb_price_text(
    price_raw: str,
    *,
    name: str = "",
    spec: str = "-",
    usage: str = "",
    role: str = "",
) -> str:
    """价格库入库保留 0.5/M、2.5/PCS、12/Y 等原始口径；仅对裸数字做展示补全。"""
    text = str(price_raw or "").strip()
    if not text or text in {"-", "—", "/"}:
        return text
    if _KB_SLASH_PRICE_RE.fullmatch(text) or _KB_YUAN_UNIT_PRICE_RE.search(text):
        return text
    if re.fullmatch(r"(\d+(?:\.\d+)?)\s*", text.replace(",", "")):
        return format_material_unit_price_text(
            text,
            name=name,
            spec=spec,
            usage=usage,
            role=role,
        )
    return text

ACTIVE_STATUSES = {"active", "启用", "有效", "正常", "使用中", ""}
INACTIVE_STATUSES = {"inactive", "停用", "disabled", "禁用"}
PENDING_STATUSES = {"pending", "待补充", "待补价", "待确认"}
AUTO_SYNC_MARKER = "AUTO_QUOTE_SYNC"
AUTO_PENDING_MARKER = "AUTO_PENDING_PRICE"
AUTO_CONFLICT_MARKER = "AUTO_PRICE_CONFLICT"
DEFAULT_HEADERS = ["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"]


@dataclass(frozen=True)
class PriceAdminEntry:
    row_id: str
    row_no: int
    name: str
    spec: str
    price: str
    marker: str
    status: str
    note: str
    updated_at: str
    updated_by: str


def list_price_entries(
    *,
    page: int = 1,
    page_size: int = 50,
    search_q: str | None = None,
    status: str | None = None,
    kb_path: Path | None = None,
) -> tuple[list[dict[str, Any]], int]:
    entries = _read_entries(_kb_target(kb_path))
    q = str(search_q or "").strip().lower()
    if q:
        entries = [
            x for x in entries if q in f"{x.name} {x.spec} {x.price} {x.note} {x.updated_by}".lower()
        ]
    sf = _normalize_status_filter(status)
    if sf == "active":
        entries = [x for x in entries if _is_active_status(x.status)]
    elif sf == "inactive":
        entries = [x for x in entries if not _is_active_status(x.status)]
    elif sf == "pending":
        entries = [x for x in entries if _is_pending_status(x.status)]
    entries.sort(key=lambda x: (x.updated_at or "", x.row_no), reverse=True)
    total = len(entries)
    p = max(1, int(page))
    ps = max(1, min(int(page_size), 200))
    start = (p - 1) * ps
    sliced = entries[start : start + ps]
    return [_entry_to_dict(x) for x in sliced], total


def price_admin_stats(kb_path: Path | None = None) -> dict[str, Any]:
    target = _kb_target(kb_path)
    if not target.is_file():
        out = {
            "total_entries": 0,
            "active_entries": 0,
            "inactive_entries": 0,
            "pending_entries": 0,
            "latest_updated_at": "",
            "kb_missing": True,
            "kb_error": f"正式价格库不存在：{target}",
        }
        out.update(admin_price_meta())
        out["quote_sync_suggestions_pending"] = _count_quote_sync_suggestions()
        out["auto_inserted_total"] = _count_auto_learn_log_entries()
        out["official_count"] = 0
        return out
    entries = _read_entries(target)
    active = [x for x in entries if _is_active_status(x.status)]
    inactive = [x for x in entries if not _is_active_status(x.status)]
    pending = [x for x in entries if _is_pending_status(x.status)]
    latest = ""
    if entries:
        latest = max((x.updated_at or "" for x in entries), default="")
    out = {
        "total_entries": len(entries),
        "active_entries": len(active),
        "inactive_entries": len(inactive),
        "pending_entries": len(pending),
        "latest_updated_at": latest,
        "kb_missing": False,
    }
    out.update(admin_price_meta())
    out["quote_sync_suggestions_pending"] = _count_quote_sync_suggestions()
    out["auto_inserted_total"] = _count_auto_learn_log_entries()
    out["official_count"] = out.get("active_entries", 0)
    return out


def price_exception_stats(exception_path: Path | None = None) -> dict[str, Any]:
    if exception_path is not None:
        entries, hidden_test = filter_visible_exceptions(_read_exception_entries(_exc_target(exception_path)))
    else:
        entries, hidden_test = filter_visible_exceptions(_load_merged_exception_entries())
    open_items = [x for x in entries if str(x.get("exception_status") or "open") == "open"]
    resolved = [x for x in entries if str(x.get("exception_status") or "") == "resolved"]
    excluded = [x for x in entries if str(x.get("exception_status") or "") == "excluded"]
    fixable = [x for x in open_items if str(x.get("review_hint") or "") == "fixable"]
    exclude_suggest = [x for x in open_items if str(x.get("review_hint") or "") == "exclude_suggest"]
    latest = max((str(x.get("updated_at") or "") for x in entries), default="")
    drop_log_count = _count_drop_log_entries()
    sugg_pending = _count_quote_sync_suggestions()
    return {
        "total_exceptions": len(entries),
        "open_exceptions": len(open_items),
        "pending_review_count": len(open_items) + sugg_pending,
        "resolved_exceptions": len(resolved),
        "excluded_exceptions": len(excluded),
        "fixable_exceptions": len(fixable),
        "exclude_suggest_exceptions": len(exclude_suggest),
        "auto_dropped_total": drop_log_count,
        "ignored_count": drop_log_count,
        "latest_updated_at": latest,
        "hidden_test_pollution": hidden_test,
        "exception_queue_is_local_cache": True,
        "exception_queue_path": str(review_exception_file()),
        "quote_sync_suggestions_pending": sugg_pending,
    }


def list_price_exceptions(
    *,
    page: int = 1,
    page_size: int = 50,
    search_q: str | None = None,
    status: str | None = "open",
    exception_path: Path | None = None,
) -> tuple[list[dict[str, Any]], int]:
    if exception_path is not None:
        entries, _hidden = filter_visible_exceptions(_read_exception_entries(_exc_target(exception_path)))
    else:
        entries, _hidden = filter_visible_exceptions(_load_merged_exception_entries())
    q = str(search_q or "").strip().lower()
    if q:
        entries = [
            x
            for x in entries
            if q
            in f"{x.get('name')} {x.get('spec')} {x.get('price')} {x.get('note')} {x.get('updated_by')}".lower()
        ]
    sf = str(status or "open").strip().lower()
    if sf in {"open", "resolved", "excluded"}:
        entries = [x for x in entries if str(x.get("exception_status") or "open") == sf]
    entries.sort(key=lambda x: (str(x.get("updated_at") or ""), str(x.get("exception_id") or "")), reverse=True)
    total = len(entries)
    p = max(1, int(page))
    ps = max(1, min(int(page_size), 200))
    start = (p - 1) * ps
    return entries[start : start + ps], total


def approve_price_exception(
    exception_id: str,
    payload: dict[str, Any],
    *,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    exception_path: Path | None = None,
) -> dict[str, Any]:
    eid = str(exception_id or payload.get("exception_id") or "").strip()
    if not eid:
        raise ValueError("缺少异常数据 ID。")
    if exception_path is not None:
        merged_list = _read_exception_entries(_exc_target(exception_path))
    else:
        merged_list = _load_merged_exception_entries()
    target = next((x for x in merged_list if str(x.get("exception_id") or "") == eid), None)
    if target is None:
        raise ValueError("异常数据不存在，请刷新后重试。")
    if str(target.get("exception_status") or "open") == "resolved":
        raise ValueError("这条异常数据已经处理过。")

    merged = dict(target)
    merged.update(payload or {})
    merged["row_id"] = ""
    merged["status"] = "active"
    merged["marker"] = str(merged.get("marker") or "").strip()
    merged["updated_by"] = str(merged.get("updated_by") or "admin")
    result = upsert_price_entry(merged, kb_path=kb_path, history_path=history_path)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sources: list[Path] = []
    if exception_path is not None:
        sources.append(_exc_target(exception_path))
    else:
        sources.extend([review_exception_file(), LEGACY_EXCEPTION_PATH])
    for src in sources:
        if not src.is_file():
            continue
        entries = _read_exception_entries(src)
        touched = False
        for item in entries:
            if str(item.get("exception_id") or "") != eid:
                continue
            item["exception_status"] = "resolved"
            item["resolved_at"] = now
            item["resolved_by"] = str(merged.get("updated_by") or "admin")
            item["approved_entry"] = result.get("entry")
            touched = True
        if touched:
            _write_exception_entries(entries, src)
    return {"ok": True, "entry": result.get("entry"), "exception_id": eid}


def delete_price_exception(
    exception_id: str,
    *,
    exception_path: Path | None = None,
    updated_by: str = "admin",
) -> dict[str, Any]:
    eid = str(exception_id or "").strip()
    if not eid:
        raise ValueError("缺少异常数据 ID。")
    path = _exc_target(exception_path)
    entries = _read_exception_entries(path)
    target = next((x for x in entries if str(x.get("exception_id") or "") == eid), None)
    if target is None:
        raise ValueError("异常数据不存在，请刷新后重试。")
    remaining = [x for x in entries if str(x.get("exception_id") or "") != eid]
    _write_exception_entries(remaining, path)
    return {
        "ok": True,
        "deleted": {
            "exception_id": eid,
            "name": str(target.get("name") or ""),
            "spec": str(target.get("spec") or ""),
            "updated_by": str(updated_by or "admin"),
        },
    }


def delete_price_exceptions_bulk(
    exception_ids: list[str],
    *,
    exception_path: Path | None = None,
    updated_by: str = "admin",
) -> dict[str, Any]:
    ids = [str(x).strip() for x in (exception_ids or []) if str(x).strip()]
    if not ids:
        raise ValueError("请至少选择一条异常数据。")
    id_set = set(ids)
    path = _exc_target(exception_path)
    entries = _read_exception_entries(path)
    deleted: list[dict[str, str]] = []
    remaining: list[dict[str, Any]] = []
    for item in entries:
        eid = str(item.get("exception_id") or "")
        if eid in id_set:
            deleted.append(
                {
                    "exception_id": eid,
                    "name": str(item.get("name") or ""),
                    "spec": str(item.get("spec") or ""),
                }
            )
            id_set.discard(eid)
        else:
            remaining.append(item)
    if not deleted:
        raise ValueError("所选异常数据不存在，请刷新后重试。")
    _write_exception_entries(remaining, path)
    return {
        "ok": True,
        "deleted_count": len(deleted),
        "deleted": deleted,
        "not_found_ids": sorted(id_set),
        "updated_by": str(updated_by or "admin"),
    }


def exclude_price_exception(
    exception_id: str,
    *,
    exception_path: Path | None = None,
    drop_log_path: Path | None = None,
    updated_by: str = "admin",
    note: str = "",
) -> dict[str, Any]:
    """人工标记排除：移出待处理队列，写入排除日志，不入正式价格库。"""
    eid = str(exception_id or "").strip()
    if not eid:
        raise ValueError("缺少异常数据 ID。")
    path = _exc_target(exception_path)
    entries = _read_exception_entries(path)
    target = next((x for x in entries if str(x.get("exception_id") or "") == eid), None)
    if target is None:
        raise ValueError("异常数据不存在，请刷新后重试。")
    if str(target.get("exception_status") or "open") != "open":
        raise ValueError("这条异常数据已经处理过。")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _append_auto_drop_log(
        {
            "ts": now,
            "action": "manual_exclude",
            "name": str(target.get("name") or ""),
            "spec": str(target.get("spec") or ""),
            "price": str(target.get("price") or ""),
            "reason": str(target.get("exception_reason") or "人工标记排除"),
            "exception_id": eid,
            "source_quote_id": str(target.get("source_quote_id") or ""),
            "updated_by": str(updated_by or "admin"),
            "note": str(note or target.get("note") or "").strip(),
        },
        _drop_target(drop_log_path),
    )
    for item in entries:
        if str(item.get("exception_id") or "") != eid:
            continue
        item["exception_status"] = "excluded"
        item["excluded_at"] = now
        item["excluded_by"] = str(updated_by or "admin")
        if note:
            item["note"] = str(note).strip()
    _write_exception_entries(entries, path)
    return {
        "ok": True,
        "exception_id": eid,
        "excluded": {
            "exception_id": eid,
            "name": str(target.get("name") or ""),
            "spec": str(target.get("spec") or ""),
        },
    }


def list_price_history(*, limit: int = 50, history_path: Path | None = None) -> list[dict[str, Any]]:
    path = _hist_target(history_path)
    if not path.exists():
        return []
    out: list[dict[str, Any]] = []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return []
    for line in reversed(lines):
        text = line.strip()
        if not text:
            continue
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            continue
        if isinstance(data, dict):
            out.append(data)
        if len(out) >= max(1, min(int(limit), 200)):
            break
    return out


def export_price_kb_workbook(
    *,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    updated_by: str = "admin",
) -> tuple[bytes, str, int]:
    """导出当前价格库为 xlsx（材料名称/规格/单价等列），由内存条目重建，避免误传 HTML。"""
    target = _kb_target(kb_path)
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = f"price_kb_{ts}.xlsx"

    if not target.exists():
        raise ValueError("价格库文件不存在，请先在后台导入 price_kb.xlsx。")
    try:
        entries = _read_entries(target)
    except Exception as exc:
        raise ValueError(f"价格库文件无法解析：{exc}") from exc
    if not entries:
        raise ValueError("价格库无有效价格行，无法导出。")

    blob = _build_workbook_bytes(entries)
    if not blob or not blob.startswith(b"PK"):
        raise RuntimeError("导出 Excel 生成失败，请检查 openpyxl 是否可用。")

    _append_history(
        {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": "export_workbook",
            "filename": filename,
            "rows": len(entries),
            "updated_by": str(updated_by or "admin"),
        },
        _hist_target(history_path),
    )
    return blob, filename, len(entries)


def import_price_kb_workbook(
    *,
    filename: str,
    content_base64: str,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    updated_by: str = "admin",
) -> dict[str, Any]:
    raw_name = str(filename or "").strip() or "import.xlsx"
    b64 = str(content_base64 or "").strip()
    if not b64:
        raise ValueError("缺少导入文件内容。")
    target = _kb_target(kb_path)
    hist = _hist_target(history_path)
    try:
        blob = base64.b64decode(b64)
    except Exception as exc:
        raise ValueError("导入文件解码失败。") from exc
    if not blob:
        raise ValueError("导入文件为空。")

    assert_official_kb_write_allowed(target, updated_by=updated_by, source="admin_import")

    tmp_path = target.with_suffix(f".import-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    bak_path = target.with_suffix(f".bak-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    tmp_path.write_bytes(blob)
    try:
        kb = get_price_kb(tmp_path)
    except Exception as exc:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise ValueError(f"导入文件格式不兼容：{exc}") from exc
    if kb.size <= 0:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise ValueError("导入文件未解析出有效价格行。")

    with KNOWLEDGE_MUTATION_LOCK:
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.exists():
            target.replace(bak_path)
        tmp_path.replace(target)
        note_kb_disk_write_success(target)
        knowledge_reload_hook(target)

    _append_history(
        {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": "import_workbook",
            "filename": raw_name,
            "new_rows": int(kb.size),
            "backup_file": str(bak_path.name if bak_path.exists() else ""),
            "updated_by": str(updated_by or "admin"),
        },
        hist,
    )
    return {
        "ok": True,
        "rows": int(kb.size),
        "backup_file": str(bak_path.name if bak_path.exists() else ""),
    }


def upsert_price_entry(
    payload: dict[str, Any],
    *,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    source: str = "admin_upsert",
) -> dict[str, Any]:
    path = _kb_target(kb_path)
    hist_path = _hist_target(history_path)
    data = _normalize_upsert_payload(payload)
    if source == "quote_auto_learn":
        data["updated_by"] = data.get("updated_by") or "quote_auto_learn"
    assert_official_kb_write_allowed(path, updated_by=data["updated_by"], source=source)
    if not data["name"]:
        raise ValueError("材料名称不能为空。")
    if not data["price"] and data["status"] != "pending":
        raise ValueError("单价不能为空；如果暂时没有单价，请先保存为待补充。")

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("需要安装 openpyxl 才能维护价格库。") from exc

    with KNOWLEDGE_MUTATION_LOCK:
        if path.exists():
            wb = load_workbook(path)
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
        try:
            ws = _pick_material_sheet(wb)
            _ensure_headers(ws)
            target_row_no = _decode_row_no(data["row_id"])
            previous = _read_entry_from_sheet_row(ws, target_row_no) if target_row_no else None
            if target_row_no and previous is None:
                raise ValueError("要编辑的价格行不存在，请刷新后重试。")
            if previous is None:
                row_no = max(ws.max_row + 1, 2)
            else:
                row_no = previous.row_no
            ws.cell(row=row_no, column=1).value = data["name"]
            ws.cell(row=row_no, column=2).value = data["spec"]
            ws.cell(row=row_no, column=3).value = data["price"]
            ws.cell(row=row_no, column=4).value = data["marker"] or (previous.marker if previous else "")
            ws.cell(row=row_no, column=5).value = data["status"]
            ws.cell(row=row_no, column=6).value = data["note"]
            ws.cell(row=row_no, column=7).value = data["updated_at"]
            ws.cell(row=row_no, column=8).value = data["updated_by"]
            wb.save(path)
            note_kb_disk_write_success(path)
            knowledge_reload_hook(path)
            current = _read_entry_from_sheet_row(ws, row_no)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    _append_history(
        {
            "ts": data["updated_at"],
            "action": "update" if previous else "create",
            "row_id": current.row_id if current else _row_id_from_no(row_no),
            "name": data["name"],
            "spec": data["spec"],
            "old_price": previous.price if previous else "",
            "new_price": data["price"],
            "old_status": previous.status if previous else "",
            "new_status": data["status"],
            "updated_by": data["updated_by"],
            "note": data["note"],
        },
        hist_path,
    )
    if current is None:
        raise RuntimeError("价格保存成功，但回读失败。")
    return {"ok": True, "entry": _entry_to_dict(current)}


def _upsert_price_entries_bulk(
    payloads: list[dict[str, Any]],
    *,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    source: str = "admin_upsert",
) -> list[dict[str, Any]]:
    if not payloads:
        return []

    rows = [_normalize_upsert_payload(payload) for payload in payloads]
    for data in rows:
        if not data["name"]:
            raise ValueError("材料名称不能为空。")
        if not data["price"] and data["status"] != "pending":
            raise ValueError("单价不能为空；如果暂时没有单价，请先保存为待补充。")

    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("需要安装 openpyxl 才能维护价格库。") from exc

    path = _kb_target(kb_path)
    hist_path = _hist_target(history_path)
    writer = str(rows[0].get("updated_by") if rows else "admin")
    if source == "quote_auto_learn":
        writer = "quote_auto_learn"
    assert_official_kb_write_allowed(path, updated_by=writer, source=source)
    saved_entries: list[PriceAdminEntry] = []
    history_records: list[dict[str, Any]] = []
    with KNOWLEDGE_MUTATION_LOCK:
        if path.exists():
            wb = load_workbook(path)
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
        try:
            ws = _pick_material_sheet(wb)
            _ensure_headers(ws)
            for data in rows:
                target_row_no = _decode_row_no(data["row_id"])
                previous = _read_entry_from_sheet_row(ws, target_row_no) if target_row_no else None
                if target_row_no and previous is None:
                    raise ValueError("要编辑的价格行不存在，请刷新后重试。")
                row_no = previous.row_no if previous is not None else max(ws.max_row + 1, 2)
                ws.cell(row=row_no, column=1).value = data["name"]
                ws.cell(row=row_no, column=2).value = data["spec"]
                ws.cell(row=row_no, column=3).value = data["price"]
                ws.cell(row=row_no, column=4).value = data["marker"] or (previous.marker if previous else "")
                ws.cell(row=row_no, column=5).value = data["status"]
                ws.cell(row=row_no, column=6).value = data["note"]
                ws.cell(row=row_no, column=7).value = data["updated_at"]
                ws.cell(row=row_no, column=8).value = data["updated_by"]
                current = _read_entry_from_sheet_row(ws, row_no)
                if current is None:
                    raise RuntimeError("价格保存成功，但回读失败。")
                saved_entries.append(current)
                history_records.append(
                    {
                        "ts": data["updated_at"],
                        "action": "update" if previous else "create",
                        "row_id": current.row_id,
                        "name": data["name"],
                        "spec": data["spec"],
                        "old_price": previous.price if previous else "",
                        "new_price": data["price"],
                        "old_status": previous.status if previous else "",
                        "new_status": data["status"],
                        "updated_by": data["updated_by"],
                        "note": data["note"],
                    }
                )
            wb.save(path)
            note_kb_disk_write_success(path)
            knowledge_reload_hook(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    for record in history_records:
        _append_history(record, hist_path)
    return [_entry_to_dict(entry) for entry in saved_entries]


def delete_price_entry(
    row_id: str,
    *,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    updated_by: str = "admin",
    name: str = "",
    spec: str = "",
    price: str = "",
) -> dict[str, Any]:
    path = _kb_target(kb_path)
    hist_path = _hist_target(history_path)
    assert_official_kb_write_allowed(path, updated_by=updated_by, source="admin_delete")
    if not path.exists():
        raise ValueError("价格库文件不存在。")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("需要安装 openpyxl 才能维护价格库。") from exc

    previous: PriceAdminEntry | None = None
    target_row_no: int | None = None
    with KNOWLEDGE_MUTATION_LOCK:
        wb = load_workbook(path)
        try:
            ws = _pick_material_sheet(wb)
            target_row_no, previous = _resolve_sheet_row_for_delete(
                ws,
                row_id=row_id,
                name=name,
                spec=spec,
                price=price,
            )
            if previous is None or target_row_no is None:
                raise ValueError("要删除的价格行不存在，请刷新后重试。")
            ws.delete_rows(target_row_no, 1)
            wb.save(path)
            note_kb_disk_write_success(path)
            knowledge_reload_hook(path)
        finally:
            try:
                wb.close()
            except Exception:
                pass

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _append_history(
        {
            "ts": now,
            "action": "delete",
            "row_id": str(row_id),
            "name": previous.name,
            "spec": previous.spec,
            "old_price": previous.price,
            "old_status": previous.status,
            "updated_by": str(updated_by or "admin"),
        },
        hist_path,
    )
    return {"ok": True, "row_id": str(row_id), "deleted": _entry_to_dict(previous)}


def _normalize_upsert_payload(payload: dict[str, Any]) -> dict[str, str]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status = _normalize_status_value(payload.get("status"))
    price_raw = str(payload.get("price") or "").strip()
    if price_raw and str(payload.get("status") or "").strip().lower() not in PENDING_STATUSES:
        price_raw = _preserve_kb_price_text(
            price_raw,
            name=str(payload.get("name") or "").strip(),
            spec=str(payload.get("spec") or "").strip() or "-",
        )
    return {
        "row_id": str(payload.get("row_id") or "").strip(),
        "name": str(payload.get("name") or "").strip(),
        "spec": str(payload.get("spec") or "").strip() or "-",
        "price": price_raw,
        "marker": str(payload.get("marker") or "").strip(),
        "status": status,
        "note": str(payload.get("note") or "").strip(),
        "updated_at": str(payload.get("updated_at") or "").strip() or now,
        "updated_by": str(payload.get("updated_by") or "").strip() or "admin",
    }


def _read_entries(path: Path) -> list[PriceAdminEntry]:
    """用 openpyxl 读取，保证 row_no 与删改行号一致（避免 XML 稀疏行错位）。"""
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _read_entries_from_sheet_xml(path)

    out: list[PriceAdminEntry] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _pick_material_sheet(wb)
        # read_only worksheets are very slow when accessed repeatedly through
        # ws.cell(row, col). Stream rows once so the admin page can load large
        # knowledge bases without timing out.
        for row_no, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            ent = _read_entry_from_values(row, row_no)
            if ent is not None:
                out.append(ent)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def _read_entry_from_values(row: Any, row_no: int) -> PriceAdminEntry | None:
    cells = list(row or ())[:8]
    while len(cells) < 8:
        cells.append("")
    name = str(cells[0] or "").strip()
    spec = str(cells[1] or "").strip()
    price = str(cells[2] or "").strip()
    if not name and not price and not spec:
        return None
    return PriceAdminEntry(
        row_id=_row_id_from_no(row_no),
        row_no=row_no,
        name=name,
        spec=spec,
        price=price,
        marker=str(cells[3] or "").strip(),
        status=_normalize_status_value(cells[4]),
        note=str(cells[5] or "").strip(),
        updated_at=str(cells[6] or "").strip(),
        updated_by=str(cells[7] or "").strip(),
    )


def _read_entries_from_sheet_xml(path: Path) -> list[PriceAdminEntry]:
    import io
    import zipfile

    with zipfile.ZipFile(io.BytesIO(path.read_bytes())) as archive:
        shared_strings = read_shared_strings(archive)
        sheets = read_sheet_entries(archive)
        if not sheets:
            return []
        _sheet_name, sheet_xml = sheets[0]
        for name, xml in sheets:
            if "询价" in name or "材料" in name:
                _sheet_name, sheet_xml = name, xml
                break
        rows = normalize_rows(parse_sheet_xml_rows(sheet_xml, shared_strings))
    out: list[PriceAdminEntry] = []
    for idx, row in enumerate(rows[1:], start=2):
        cells = [str(c or "").strip() for c in row[:8]]
        while len(cells) < 8:
            cells.append("")
        if not any(cells[:3]):
            continue
        out.append(
            PriceAdminEntry(
                row_id=_row_id_from_no(idx),
                row_no=idx,
                name=cells[0],
                spec=cells[1],
                price=cells[2],
                marker=cells[3],
                status=_normalize_status_value(cells[4]),
                note=cells[5],
                updated_at=cells[6],
                updated_by=cells[7],
            )
        )
    return out


def _resolve_sheet_row_for_delete(
    ws,
    *,
    row_id: str,
    name: str = "",
    spec: str = "",
    price: str = "",
) -> tuple[int | None, PriceAdminEntry | None]:
    hint_row = _decode_row_no(row_id)
    want_name = str(name or "").strip()
    want_spec = str(spec or "").strip() or "-"
    want_price = str(price or "").strip()
    norm_name = _norm_key(want_name)
    norm_spec = _norm_key(want_spec)
    norm_price = _norm_price(want_price)

    if hint_row is not None:
        ent = _read_entry_from_sheet_row(ws, hint_row)
        if ent is not None and (
            not norm_name
            or (
                _norm_key(ent.name) == norm_name
                and _norm_key(ent.spec) == norm_spec
                and (not norm_price or _norm_price(ent.price) == norm_price)
            )
        ):
            return hint_row, ent

    max_row = int(getattr(ws, "max_row", 0) or 0)
    for row_no in range(2, max_row + 1):
        ent = _read_entry_from_sheet_row(ws, row_no)
        if ent is None:
            continue
        if norm_name and (_norm_key(ent.name) != norm_name or _norm_key(ent.spec) != norm_spec):
            continue
        if norm_price and _norm_price(ent.price) != norm_price:
            continue
        if not norm_name and hint_row is not None and row_no != hint_row:
            continue
        return row_no, ent
    return None, None


def _pick_material_sheet(wb):
    for ws in wb.worksheets:
        name = str(ws.title or "")
        if "询价" in name or "材料" in name:
            return ws
    return wb.worksheets[0]


def _ensure_headers(ws) -> None:
    for idx, title in enumerate(DEFAULT_HEADERS, start=1):
        if not ws.cell(row=1, column=idx).value:
            ws.cell(row=1, column=idx).value = title


def _build_workbook_bytes(entries: list[PriceAdminEntry]) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("需要安装 openpyxl 才能导出价格库。") from exc

    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "材料询价"
    _ensure_headers(ws)
    ordered = sorted(entries, key=lambda x: x.row_no)
    for entry in ordered:
        ws.append(
            [
                entry.name,
                entry.spec,
                entry.price,
                entry.marker,
                _status_export_label(entry.status),
                entry.note,
                entry.updated_at,
                entry.updated_by,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    try:
        wb.close()
    except Exception:
        pass
    return buf.getvalue()


def _status_export_label(status: str) -> str:
    norm = _normalize_status_value(status)
    if norm == "active":
        return "启用"
    if norm == "inactive":
        return "停用"
    if norm == "pending":
        return "待补充"
    return str(status or "").strip() or "启用"


def _read_entry_from_sheet_row(ws, row_no: int | None) -> PriceAdminEntry | None:
    if row_no is None or row_no < 2 or row_no > ws.max_row:
        return None
    cells = [ws.cell(row=row_no, column=i).value for i in range(1, 9)]
    name = str(cells[0] or "").strip()
    spec = str(cells[1] or "").strip()
    price = str(cells[2] or "").strip()
    if not name and not price and not spec:
        return None
    return PriceAdminEntry(
        row_id=_row_id_from_no(row_no),
        row_no=row_no,
        name=name,
        spec=spec,
        price=price,
        marker=str(cells[3] or "").strip(),
        status=_normalize_status_value(cells[4]),
        note=str(cells[5] or "").strip(),
        updated_at=str(cells[6] or "").strip(),
        updated_by=str(cells[7] or "").strip(),
    )


def _row_id_from_no(row_no: int) -> str:
    return f"row-{int(row_no)}"


def _decode_row_no(row_id: str) -> int | None:
    text = str(row_id or "").strip().lower()
    if not text.startswith("row-"):
        return None
    try:
        value = int(text.split("-", 1)[1])
    except (TypeError, ValueError):
        return None
    return value if value >= 2 else None


def _normalize_status_filter(value: str | None) -> str:
    text = str(value or "").strip().lower()
    if text in {"active", "inactive", "pending"}:
        return text
    return ""


def _normalize_status_value(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return "active"
    low = text.lower()
    if low in ACTIVE_STATUSES:
        return "active"
    if low in INACTIVE_STATUSES:
        return "inactive"
    if low in PENDING_STATUSES:
        return "pending"
    return text


def _is_active_status(value: object) -> bool:
    return _normalize_status_value(value) == "active"


def _is_pending_status(value: object) -> bool:
    return _normalize_status_value(value) == "pending"


def _entry_to_dict(entry: PriceAdminEntry) -> dict[str, Any]:
    return {
        "row_id": entry.row_id,
        "row_no": entry.row_no,
        "name": entry.name,
        "spec": entry.spec,
        "price": entry.price,
        "marker": entry.marker,
        "status": entry.status,
        "note": entry.note,
        "updated_at": entry.updated_at,
        "updated_by": entry.updated_by,
        "is_active": _is_active_status(entry.status),
    }


def _append_history(record: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def _read_exception_entries(path: Path) -> list[dict[str, Any]]:
    path = path.resolve()
    if not path.exists():
        return []
    out: list[dict[str, Any]] = []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return []
    for line in lines:
        text = line.strip()
        if not text:
            continue
        try:
            row = json.loads(text)
        except json.JSONDecodeError:
            continue
        if isinstance(row, dict):
            out.append(_normalize_exception_row(row))
    return out


def _write_exception_entries(entries: list[dict[str, Any]], path: Path) -> None:
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    text = "\n".join(json.dumps(_normalize_exception_row(x), ensure_ascii=False) for x in entries)
    path.write_text((text + "\n") if text else "", encoding="utf-8")


def _append_price_exceptions(rows: list[dict[str, Any]], path: Path) -> list[dict[str, Any]]:
    if not rows:
        return []
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    normalized = [_normalize_exception_row(x) for x in rows]
    with path.open("a", encoding="utf-8") as fh:
        for row in normalized:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")
    return normalized


def _append_auto_drop_log(record: dict[str, Any], path: Path | None = None) -> None:
    target = _drop_target(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record, ensure_ascii=False) + "\n")


def _count_drop_log_entries(path: Path | None = None) -> int:
    target = _drop_target(path)
    if not target.exists():
        return 0
    try:
        lines = target.read_text(encoding="utf-8").splitlines()
    except OSError:
        return 0
    return sum(1 for line in lines if line.strip())


def _derive_exception_reason_from_note(note: str) -> str:
    text = str(note or "").strip()
    if "无法自动分摊" in text or "组合材料" in text:
        return "组合材料需拆分补价"
    if "AI/系统估算" in text or "AI" in text and "估算" in text:
        return "AI单价需确认"
    if "缺少单价" in text or "缺少可信单价" in text or "请补价" in text:
        return "缺少价格"
    if "单价格式" in text or "缺少单位" in text:
        return "价格格式异常"
    if "规格" in text and ("缺少" in text or "不足" in text):
        return "规格缺失"
    if "数据质量待确认" in text and "：" in text:
        fragment = text.split("数据质量待确认：", 1)[-1].split("。", 1)[0].strip()
        if fragment:
            return format_exception_reason_label(
                KbDataQualityVerdict(KB_ACTION_REVIEW, fragment, "suspicious"),
            )
    return "待人工确认"


def _derive_review_hint_from_reason(reason: str) -> str:
    text = str(reason or "").strip()
    if any(token in text for token in ("疑似非材料", "混合", "说明", "排除")):
        return "exclude_suggest"
    if any(token in text for token in ("缺少", "格式", "分摊", "AI", "规格")):
        return "fixable"
    return "review"


def _build_exception_payload(
    *,
    name: str,
    spec: str,
    price: str,
    marker: str,
    updated_by: str,
    note: str,
    quality,
    source_quote_id: str = "",
    product_name: str = "",
    is_combined_split: bool = False,
) -> dict[str, Any]:
    has_price = _has_usable_price(price)
    return {
        "name": name,
        "spec": spec or "-",
        "price": price,
        "marker": marker,
        "status": "open",
        "updated_by": updated_by,
        "note": note,
        "source_quote_id": source_quote_id,
        "product_name": product_name,
        "exception_reason": format_exception_reason_label(quality, is_combined_split=is_combined_split),
        "review_hint": classify_exception_review_hint(
            name,
            quality,
            has_price=has_price,
            is_combined_split=is_combined_split,
        ),
    }


def _normalize_exception_row(row: dict[str, Any]) -> dict[str, Any]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status = str(row.get("exception_status") or row.get("status") or "open").strip().lower()
    if status not in {"open", "resolved", "excluded"}:
        status = "open"
    eid = str(row.get("exception_id") or "").strip() or f"ex-{uuid.uuid4().hex}"
    reason = str(row.get("exception_reason") or "").strip()
    if not reason:
        reason = _derive_exception_reason_from_note(str(row.get("note") or ""))
    review_hint = str(row.get("review_hint") or "").strip()
    if review_hint not in {"fixable", "exclude_suggest", "review"}:
        review_hint = _derive_review_hint_from_reason(reason)
    return {
        "exception_id": eid,
        "row_id": eid,
        "name": str(row.get("name") or "").strip(),
        "spec": str(row.get("spec") or "").strip() or "-",
        "price": str(row.get("price") or "").strip(),
        "marker": str(row.get("marker") or "").strip(),
        "status": "pending",
        "exception_status": status,
        "exception_reason": reason,
        "review_hint": review_hint,
        "note": str(row.get("note") or "").strip(),
        "updated_at": str(row.get("updated_at") or "").strip() or now,
        "updated_by": str(row.get("updated_by") or "").strip() or "agent_auto",
        "source_quote_id": str(row.get("source_quote_id") or "").strip(),
        "product_name": str(row.get("product_name") or "").strip(),
        "is_exception": True,
    }


def _auto_insert_trusted_entries(
    payloads: list[dict[str, str]],
    *,
    kb_path: Path | None,
    history_path: Path | None,
    quote_id: str = "",
    product_name: str = "",
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    """可信材料自动写入 KB；正式库禁止时退回 suggestions 队列。"""
    if not payloads:
        return [], []
    path = _kb_target(kb_path)
    try:
        assert_official_kb_write_allowed(path, updated_by="quote_auto_learn", source="quote_auto_learn")
    except PermissionError:
        return [], payloads

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for raw in payloads:
        item = dict(raw)
        item.setdefault("updated_at", now)
        item.setdefault("updated_by", "quote_auto_learn")
        item.setdefault("status", "active")
        rows.append(item)
    try:
        saved = _upsert_price_entries_bulk(
            rows,
            kb_path=kb_path,
            history_path=history_path,
            source="quote_auto_learn",
        )
    except Exception:
        return [], payloads

    for entry in saved:
        _append_auto_learn_log(
            {
                "ts": now,
                "action": "auto_insert",
                "name": entry.get("name"),
                "spec": entry.get("spec"),
                "price": entry.get("price"),
                "source_quote_id": quote_id,
                "product_name": product_name,
            }
        )
    return saved, []


def sync_quote_detail_rows_to_price_kb(
    quote_result: dict[str, Any],
    *,
    kb_path: Path | None = None,
    history_path: Path | None = None,
    exception_path: Path | None = None,
    updated_by: str = "agent_auto",
) -> dict[str, Any]:
    """报价完成后自动学习价格库。

    规则：
    - 名称 + 规格 + 单价完全相同：跳过；
    - 名称 + 规格相同但单价不同：进待审核异常队列；
    - 可信新材料且有单价：自动写入 KB（正式库需 ALLOW_OFFICIAL_KB_AUTO_LEARN）；
    - 可疑/缺价：进待审核异常队列；
    - 非材料/垃圾：丢弃并记日志。
    """
    if not isinstance(quote_result, dict):
        return _sync_summary()
    if is_test_quote_sync_context(quote_result):
        summary = _sync_summary()
        summary["skipped"] += 1
        summary["ignored_test_quote"] = True
        return summary
    intent = str(quote_result.get("intent") or quote_result.get("flow_intent") or "").strip()
    if intent in {"agent_trial", "extra_material_calc", "extra_quantity_calc"}:
        return _sync_summary()
    md = quote_result.get("metadata") if isinstance(quote_result.get("metadata"), dict) else {}
    if md.get("is_extra_calc") or md.get("is_extra_material_calc"):
        return _sync_summary()
    rows = quote_result.get("detail_rows")
    if not isinstance(rows, list):
        return _sync_summary()

    existing = [_entry_to_dict(x) for x in _read_entries(_kb_target(kb_path))]
    by_key: dict[tuple[str, str], list[dict[str, Any]]] = {}
    exact: set[tuple[str, str, str]] = set()
    for item in existing:
        key = (_norm_key(item.get("name")), _norm_key(item.get("spec")))
        price_key = _norm_price(item.get("price"))
        by_key.setdefault(key, []).append(item)
        exact.add((key[0], key[1], price_key))

    summary = _sync_summary()
    seen_this_quote: set[tuple[str, str, str]] = set()
    quote_id = str(quote_result.get("quote_id") or "").strip()
    product_name = str(quote_result.get("product_name") or "").strip()
    payloads_to_save: list[dict[str, str]] = []
    exceptions_to_save: list[dict[str, Any]] = []

    for row in rows:
        if not isinstance(row, dict):
            continue
        rec_status = str(row.get("recognition_status") or "").strip()
        if rec_status == "ignored":
            summary["dropped"] = int(summary.get("dropped") or 0) + 1
            summary["ignored"] = int(summary.get("ignored") or 0) + 1
            summary["skipped"] += 1
            continue
        raw_name = str(row.get("name") or "").strip()
        if not _valid_material_name(raw_name):
            continue
        split_names = _split_combined_material_names(_kb_material_name_from_quote_name(raw_name))
        is_combined_name = len(split_names) > 1
        spec = str(row.get("spec") or "").strip() or "-"
        row_price = str(row.get("unit_price") or "").strip()
        row_role = str(row.get("role") or "").strip()
        row_price = _preserve_kb_price_text(
            row_price,
            name=raw_name,
            spec=spec,
            usage=_row_usage_text(row),
            role=row_role,
        )
        if _has_kb_garbage_symbol(raw_name, spec, row_price):
            summary["skipped"] += 1
            continue
        row_usage = _row_usage_text(row)
        if not _has_usable_price(row_price) and not _has_meaningful_spec(spec) and not _has_meaningful_usage(row_usage):
            summary["skipped"] += 1
            continue
        row_price_needs_review = _row_price_needs_human_review(row)

        for name in split_names:
            if not _valid_material_name(name):
                continue
            price = "" if is_combined_name else row_price
            if _has_kb_garbage_symbol(name, spec, price):
                summary["dropped"] = int(summary.get("dropped") or 0) + 1
                summary["skipped"] += 1
                continue

            quality = judge_kb_insert_candidate(
                name,
                spec,
                price,
                row=row,
                kb_hit=bool(row.get("kb_hit")),
            )
            if quality.action == KB_ACTION_DROP:
                summary["dropped"] = int(summary.get("dropped") or 0) + 1
                summary["ignored"] = int(summary.get("ignored") or 0) + 1
                summary["skipped"] += 1
                _append_auto_drop_log(
                    {
                        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "action": "auto_drop",
                        "name": name,
                        "spec": spec,
                        "price": price,
                        "reason": quality.reason,
                        "source_quote_id": quote_id,
                        "product_name": product_name,
                    },
                )
                continue

            key = (_norm_key(name), _norm_key(spec))
            price_key = _norm_price(price)
            if not key[0]:
                continue
            quote_seen_key = (key[0], key[1], price_key)
            if quote_seen_key in seen_this_quote:
                summary["skipped"] += 1
                continue
            seen_this_quote.add(quote_seen_key)

            if price_key and (key[0], key[1], price_key) in exact:
                summary["skipped"] += 1
                continue

            same_key_rows = by_key.get(key, [])
            has_price = _has_usable_price(price)
            force_review = quality.action == KB_ACTION_REVIEW or row_price_needs_review
            if same_key_rows:
                if has_price and price_key:
                    existing_prices = {
                        _norm_price(item.get("price"))
                        for item in same_key_rows
                        if _norm_price(item.get("price"))
                    }
                    if price_key in existing_prices:
                        summary["skipped"] += 1
                        continue
                    note = (
                        f"同材料同规格价格冲突：知识库 {', '.join(sorted(existing_prices))}，"
                        f"报价 {price}；需人工确认。"
                        f"报价ID：{quote_id}；产品：{product_name}"
                    )
                    payload = _build_exception_payload(
                        name=name,
                        spec=spec,
                        price=price,
                        marker=AUTO_CONFLICT_MARKER,
                        updated_by=updated_by,
                        note=note,
                        quality=KbDataQualityVerdict(
                            KB_ACTION_REVIEW,
                            "同材料同规格价格差异明显",
                            "suspicious",
                        ),
                        source_quote_id=quote_id,
                        product_name=product_name,
                    )
                    exceptions_to_save.append(payload)
                    summary["conflicts"] += 1
                    summary["pending"] += 1
                    continue
                summary["skipped"] += 1
                continue
            if has_price and force_review:
                note = (
                    f"数据质量待确认：{quality.reason}。"
                    f"报价中「{name}」需人工审核后再启用。"
                    f"报价ID：{quote_id}；产品：{product_name}"
                )
                payload = _build_exception_payload(
                    name=name,
                    spec=spec,
                    price=price,
                    marker=AUTO_PENDING_MARKER,
                    updated_by=updated_by,
                    note=note,
                    quality=quality,
                    source_quote_id=quote_id,
                    product_name=product_name,
                )
                exceptions_to_save.append(payload)
                summary["pending"] += 1
            elif has_price and quality.action == KB_ACTION_AUTO:
                payload = _auto_price_payload(
                    name=name,
                    spec=spec,
                    price=price,
                    status="active",
                    marker=AUTO_SYNC_MARKER,
                    updated_by="quote_auto_learn",
                    note=f"报价自动补入；报价ID：{quote_id}；产品：{product_name}",
                )
                payloads_to_save.append(payload)
            else:
                if is_combined_name:
                    note = (
                        f"报价发现组合材料「{raw_name}」，已拆分为「{name}」；"
                        f"原组合单价 {row_price or '-'} 无法自动分摊，请补单价后启用。"
                        f"报价ID：{quote_id}；产品：{product_name}"
                    )
                else:
                    note = (
                        f"数据质量待确认：{quality.reason}。"
                        f"报价发现新材料但缺少单价，请补价后启用。"
                        f"报价ID：{quote_id}；产品：{product_name}"
                    )
                payload = _build_exception_payload(
                    name=name,
                    spec=spec,
                    price="",
                    marker=AUTO_PENDING_MARKER,
                    updated_by=updated_by,
                    note=note,
                    quality=quality,
                    source_quote_id=quote_id,
                    product_name=product_name,
                    is_combined_split=is_combined_name,
                )
                exceptions_to_save.append(payload)
                summary["pending"] += 1

            by_key.setdefault(key, []).append(payload)
            exact.add((key[0], key[1], price_key))

    exceptions_to_save = [
        x for x in exceptions_to_save if not is_test_price_exception_record(x)
    ]
    saved_exceptions = _append_price_exceptions(exceptions_to_save, _exc_target(exception_path))
    for entry in saved_exceptions:
        summary["items"].append(
            {
                "name": entry.get("name"),
                "spec": entry.get("spec"),
                "price": entry.get("price"),
                "status": "exception",
                "marker": entry.get("marker"),
                "exception_id": entry.get("exception_id"),
            }
        )

    if payloads_to_save:
        inserted, fallback = _auto_insert_trusted_entries(
            payloads_to_save,
            kb_path=kb_path,
            history_path=history_path,
            quote_id=quote_id,
            product_name=product_name,
        )
        summary["created"] = int(summary.get("created") or 0) + len(inserted)
        summary["auto_inserted"] = int(summary.get("auto_inserted") or 0) + len(inserted)
        for entry in inserted:
            summary["items"].append(
                {
                    "name": entry.get("name"),
                    "spec": entry.get("spec"),
                    "price": entry.get("price"),
                    "status": "active",
                    "marker": AUTO_SYNC_MARKER,
                }
            )
        if fallback:
            queued = _append_quote_sync_suggestions(fallback)
            summary["suggestions_queued"] = int(summary.get("suggestions_queued") or 0) + len(queued)
            for payload in queued:
                summary["items"].append(
                    {
                        "name": payload.get("name"),
                        "spec": payload.get("spec"),
                        "price": payload.get("price"),
                        "status": "pending_review",
                        "marker": payload.get("marker"),
                    }
                )

    return summary


def _sync_summary() -> dict[str, Any]:
    return {
        "created": 0,
        "auto_inserted": 0,
        "pending": 0,
        "conflicts": 0,
        "skipped": 0,
        "dropped": 0,
        "ignored": 0,
        "suggestions_queued": 0,
        "items": [],
    }


def _auto_price_payload(
    *,
    name: str,
    spec: str,
    price: str,
    status: str,
    marker: str,
    updated_by: str,
    note: str,
) -> dict[str, str]:
    return {
        "name": name,
        "spec": spec or "-",
        "price": price,
        "marker": marker,
        "status": status,
        "updated_by": updated_by or "agent_auto",
        "note": note,
    }


def _norm_key(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"", "-", "—", "/"}:
        return ""
    return "".join(text.split())


def _norm_price(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"", "-", "—", "/"}:
        return ""
    return "".join(text.split())


def _has_usable_price(value: object) -> bool:
    text = str(value or "").strip()
    if _has_kb_garbage_symbol(text):
        return False
    return bool(_norm_price(text))


def _has_kb_garbage_symbol(*values: object) -> bool:
    return any("?" in str(value or "") or "？" in str(value or "") for value in values)


def _has_meaningful_spec(value: object) -> bool:
    return bool(_norm_key(value))


def _row_usage_text(row: dict[str, Any]) -> str:
    for key in ("usage", "use_amount", "quantity", "qty", "amount_usage"):
        text = str(row.get(key) or "").strip()
        if text:
            return text
    return ""


def _has_meaningful_usage(value: object) -> bool:
    text = str(value or "").strip()
    if _has_kb_garbage_symbol(text):
        return False
    return bool(_norm_key(text))


def _truthy(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "on", "是", "已", "ai", "true"}


def _row_price_needs_human_review(row: dict[str, Any]) -> bool:
    source = str(row.get("source") or "").strip().lower()
    if source in {"ai", "estimated", "estimate", "system", "fallback"}:
        return True
    if _truthy(row.get("unit_price_ai")) or _truthy(row.get("price_ai")):
        return True
    if _truthy(row.get("amount_ai")) and not _truthy(row.get("kb_hit")):
        return True
    if str(row.get("ai_reason") or "").strip():
        return True
    return False


_ROLE_ONLY_MATERIAL_NAMES = {
    "\u4e3b\u4f53\u9762\u6599",
    "\u4e3b\u9762\u6599",
    "\u5916\u6599",
    "\u5916\u5c42",
    "\u5185\u886c",
    "\u5185\u91cc",
    "\u91cc\u5e03",
    "\u91cc\u6599",
    "\u5e95\u90e8\u8d34\u7247",
    "\u5e95\u7247",
    "\u8d34\u7247",
}
_ROLE_SUFFIX_RE = re.compile(
    r"[\(\uff08]\s*(?:"
    r"\u4e3b\u4f53\u9762\u6599|\u4e3b\u9762\u6599|\u5916\u6599|\u5916\u5c42|"
    r"\u5185\u886c|\u5185\u91cc|\u91cc\u5e03|\u91cc\u6599|"
    r"\u5e95\u90e8\u8d34\u7247|\u5e95\u7247|\u8d34\u7247"
    r")\s*[\)\uff09]?$",
    re.I,
)
_USAGE_ONLY_NAME_RE = re.compile(
    r"^(?:\u4ec5\u7528\u4e8e|\u53ea\u7528\u4e8e|\u4e13\u7528\u4e8e|"
    r"\u7528\u4e8e|\u4f5c\u4e3a|\u4e3b\u8981\u7528\u4e8e).*(?:"
    r"\u5305\u4f53|\u5305\u8eab|\u5e95\u90e8|\u6700\u4e0b\u65b9|\u4e0b\u65b9|"
    r"\u8d34\u7247|\u5e95\u7247|\u90e8\u4f4d|\u4f4d\u7f6e|\u8865\u5f3a"
    r")",
    re.I,
)


def _kb_material_name_from_quote_name(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = _ROLE_SUFFIX_RE.sub("", text).strip()
    text = re.sub(r"[\(\uff08]\s*$", "", text).strip()
    if not text or text in _ROLE_ONLY_MATERIAL_NAMES:
        return ""
    if _USAGE_ONLY_NAME_RE.search(text):
        return ""
    return text


def _valid_material_name(value: object) -> bool:
    text = _kb_material_name_from_quote_name(value)
    if not text or text in {"-", "—", "/"}:
        return False
    blocked = ("合计", "小计", "系统成本", "加工费", "杂费", "管理费", "开模", "模具")
    blocked_estimated = (
        "系统估算",
        "包装费",
        "包装袋",
        "纸箱",
        "纸盒",
        "外箱",
        "外纸箱",
        "装箱",
        "封箱",
    )
    if any(token in text for token in blocked_estimated):
        return False
    return not any(token in text for token in blocked)


def _split_combined_material_names(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    parts = [x.strip() for x in re.split(r"[，,；;、+＋]+", text) if x.strip()]
    return parts if len(parts) > 1 else [text]
