import json
import shutil
import unittest
import uuid
from pathlib import Path

from core.knowledge_pending_apply import apply_pending_auto_learn
from price_kb import get_price_kb, reset_price_kb


try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    Workbook = None


@unittest.skipIf(Workbook is None, "openpyxl is required")
class KnowledgePendingApplyTest(unittest.TestCase):
    def test_pending_records_are_applied_and_queue_is_rewritten(self) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_pending_apply_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            pending_file = root / "pending_auto_learn.jsonl"

            wb = Workbook()
            ws = wb.active
            ws.title = "Price KB"
            ws.append(["name", "spec", "price"])
            ws.append(["existing-auto-row", "-", "1.23/PCS"])
            wb.save(kb_path)
            wb.close()

            records = [
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.96,
                    "material": {
                        "name": "auto-learn-new-row",
                        "spec": "XL",
                        "price": "2.50/PCS",
                    },
                },
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.99,
                    "material": {
                        "name": "existing-auto-row",
                        "spec": "-",
                        "price": "1.23/PCS",
                    },
                },
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.10,
                    "material": {
                        "name": "low-confidence-row",
                        "spec": "-",
                        "price": "9.99/PCS",
                    },
                },
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.99,
                    "material": {"name": "", "spec": "-", "price": ""},
                },
            ]
            pending_file.write_text(
                "\n".join(json.dumps(r, ensure_ascii=False) for r in records)
                + "\n"
                + "{bad json\n",
                encoding="utf-8",
            )

            result = apply_pending_auto_learn(
                pending_file=pending_file,
                kb_path=kb_path,
                min_confidence=0.8,
                reload_after_write=False,
            )

            self.assertEqual(result.total, 5)
            self.assertEqual(result.applied, 1)
            self.assertEqual(result.skipped_existing, 1)
            self.assertEqual(result.invalid, 3)
            self.assertEqual(result.failed, 0)
            self.assertEqual(result.kept, 3)

            reset_price_kb()
            kb = get_price_kb(kb_path)
            self.assertEqual(kb.size, 2)
            hit = kb.lookup("auto-learn-new-row", "XL", min_score=0.1)
            self.assertIsNotNone(hit)
            assert hit is not None
            self.assertTrue(hit.entry.auto_learned)

            kept_lines = [line for line in pending_file.read_text(encoding="utf-8").splitlines() if line]
            self.assertEqual(len(kept_lines), 3)
            kept = [json.loads(line) for line in kept_lines]
            reasons = {str(row.get("_pending_apply_error") or row.get("_error")) for row in kept}
            self.assertIn("low_confidence", reasons)
            self.assertIn("missing_material_name_or_price", reasons)
            self.assertIn("invalid_json", reasons)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_missing_pending_file_is_noop(self) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_pending_missing_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            result = apply_pending_auto_learn(
                pending_file=root / "missing.jsonl",
                kb_path=root / "price_kb.xlsx",
            )

            self.assertEqual(result.total, 0)
            self.assertEqual(result.applied, 0)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_pending_records_with_question_marks_are_not_applied(self) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_pending_guard_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            pending_file = root / "pending_auto_learn.jsonl"

            wb = Workbook()
            ws = wb.active
            ws.title = "Price KB"
            ws.append(["name", "spec", "price"])
            wb.save(kb_path)
            wb.close()

            records = [
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.99,
                    "material": {"name": "300D??", "spec": "152cm", "price": "10.5/Y"},
                },
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.99,
                    "material": {"name": "210D", "spec": "58#", "price": "12.5/?"},
                },
                {
                    "type": "kb_auto_learn_candidate",
                    "confidence": 0.99,
                    "material": {"name": "safe-row", "spec": "M", "price": "1.2/PCS"},
                },
            ]
            pending_file.write_text(
                "\n".join(json.dumps(r, ensure_ascii=False) for r in records) + "\n",
                encoding="utf-8",
            )

            result = apply_pending_auto_learn(
                pending_file=pending_file,
                kb_path=kb_path,
                min_confidence=0.8,
                reload_after_write=False,
            )

            self.assertEqual(result.total, 3)
            self.assertEqual(result.applied, 1)
            self.assertEqual(result.invalid, 2)
            reset_price_kb()
            kb = get_price_kb(kb_path)
            self.assertIsNone(kb.lookup("300D??", "152cm", min_score=0.1))
            self.assertIsNone(kb.lookup("210D", "58#", min_score=0.1))
            self.assertIsNotNone(kb.lookup("safe-row", "M", min_score=0.1))
            kept = [json.loads(line) for line in pending_file.read_text(encoding="utf-8").splitlines() if line]
            reasons = {str(row.get("_pending_apply_error") or "") for row in kept}
            self.assertEqual(len(reasons), 1)
            self.assertTrue(next(iter(reasons)).startswith("quality_drop:"))
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_pending_records_with_description_name_are_not_applied(self) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_pending_desc_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            pending_file = root / "pending_auto_learn.jsonl"

            wb = Workbook()
            ws = wb.active
            ws.title = "Price KB"
            ws.append(["name", "spec", "price"])
            wb.save(kb_path)
            wb.close()

            rec = {
                "type": "kb_auto_learn_candidate",
                "confidence": 0.99,
                "material": {"name": "外侧使用主面料", "spec": "-", "price": "0.6元/个"},
            }
            pending_file.write_text(json.dumps(rec, ensure_ascii=False) + "\n", encoding="utf-8")

            result = apply_pending_auto_learn(
                pending_file=pending_file,
                kb_path=kb_path,
                min_confidence=0.8,
                reload_after_write=False,
            )

            self.assertEqual(result.total, 1)
            self.assertEqual(result.applied, 0)
            self.assertEqual(result.invalid, 1)
            kept = [json.loads(line) for line in pending_file.read_text(encoding="utf-8").splitlines() if line]
            err = str(kept[0].get("_pending_apply_error") or "")
            self.assertTrue(err.startswith("quality_drop:"))
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_pending_records_skip_existing_name_spec_even_if_price_differs(self) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_pending_namespec_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            pending_file = root / "pending_auto_learn.jsonl"

            wb = Workbook()
            ws = wb.active
            ws.title = "Price KB"
            ws.append(["name", "spec", "price"])
            ws.append(["5#YKK防水拉链—国产", "5#", "7.5元/条"])
            wb.save(kb_path)
            wb.close()

            rec = {
                "type": "kb_auto_learn_candidate",
                "confidence": 0.99,
                "material": {"name": "5#YKK防水拉链—国产", "spec": "5#", "price": "8.0元/条"},
            }
            pending_file.write_text(json.dumps(rec, ensure_ascii=False) + "\n", encoding="utf-8")

            result = apply_pending_auto_learn(
                pending_file=pending_file,
                kb_path=kb_path,
                min_confidence=0.8,
                reload_after_write=False,
            )

            self.assertEqual(result.total, 1)
            self.assertEqual(result.applied, 0)
            self.assertEqual(result.skipped_existing, 1)
            self.assertEqual([line for line in pending_file.read_text(encoding="utf-8").splitlines() if line], [])
        finally:
            shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
