from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
import shutil
import uuid
import base64
from unittest.mock import patch

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None

from price_admin_store import (
    AUTO_CONFLICT_MARKER,
    AUTO_PENDING_MARKER,
    AUTO_SYNC_MARKER,
    approve_price_exception,
    delete_price_entry,
    export_price_kb_workbook,
    import_price_kb_workbook,
    list_price_exceptions,
    list_price_entries,
    list_price_history,
    sync_quote_detail_rows_to_price_kb,
    upsert_price_entry,
)
from price_kb import get_price_kb


@unittest.skipIf(Workbook is None, "openpyxl is required")
class PriceAdminStoreTest(unittest.TestCase):
    def _make_kb(self, root: Path) -> Path:
        path = root / "price_kb.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "材料询价"
        ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
        ws.append(["600D塔丝隆格子布", "150CM", "12元/码", "", "active", "初始价", "2026-05-19 09:00:00", "seed"])
        wb.save(path)
        wb.close()
        return path

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_create_and_update_price_entry(self, _mock_note: object, _mock_reload: object) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_admin_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = self._make_kb(root)
            history_path = root / "price_kb_history.jsonl"

            created = upsert_price_entry(
                {
                    "name": "210D涤纶",
                    "spec": "152CM",
                    "price": "3元/码",
                    "status": "active",
                    "updated_by": "pm",
                    "note": "新增里料价",
                },
                kb_path=kb_path,
                history_path=history_path,
            )
            self.assertTrue(created["ok"])
            self.assertEqual(created["entry"]["name"], "210D涤纶")

            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 2)
            target = next(x for x in items if x["name"] == "210D涤纶")
            self.assertEqual(target["price"], "3元/码")

            updated = upsert_price_entry(
                {
                    "row_id": target["row_id"],
                    "name": "210D涤纶",
                    "spec": "152CM",
                    "price": "3.2元/码",
                    "status": "inactive",
                    "updated_by": "pm",
                    "note": "供应商涨价",
                },
                kb_path=kb_path,
                history_path=history_path,
            )
            self.assertTrue(updated["ok"])
            self.assertEqual(updated["entry"]["status"], "inactive")

            items2, total2 = list_price_entries(
                page=1,
                page_size=20,
                search_q="210D",
                kb_path=kb_path,
            )
            self.assertEqual(total2, 1)
            self.assertEqual(items2[0]["price"], "3.2元/码")
            self.assertFalse(items2[0]["is_active"])

            history = list_price_history(limit=10, history_path=history_path)
            self.assertEqual(len(history), 2)
            self.assertEqual(history[0]["new_price"], "3.2元/码")
            self.assertEqual(history[1]["action"], "create")
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_delete_price_entry(self, _mock_note: object, _mock_reload: object) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_del_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            ws.append(["KEEP_ME", "A", "1/Y", "", "active", "", "2026-05-19 10:00:00", "tester"])
            ws.append(["DELETE_ME", "B", "2/Y", "", "active", "", "2026-05-19 10:01:00", "tester"])
            wb.save(kb_path)
            wb.close()

            out = delete_price_entry(
                "row-3",
                kb_path=kb_path,
                history_path=root / "history.jsonl",
                updated_by="tester",
            )
            self.assertTrue(out.get("ok"))
            self.assertEqual(out.get("deleted", {}).get("name"), "DELETE_ME")

            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 1)
            self.assertEqual(items[0]["name"], "KEEP_ME")

            history = list_price_history(limit=3, history_path=root / "history.jsonl")
            self.assertEqual(history[0]["action"], "delete")
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def test_export_price_workbook_reads_disk_file(self) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_export_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            target_kb = root / "price_kb.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            ws.append(["EXPORT_TEST", "150CM", "9.5元/码", "", "启用", "", "2026-05-19 10:00:00", "tester"])
            wb.save(target_kb)
            wb.close()

            blob, filename, rows = export_price_kb_workbook(
                kb_path=target_kb,
                history_path=root / "history.jsonl",
                updated_by="tester",
            )
            self.assertTrue(filename.endswith(".xlsx"))
            self.assertEqual(rows, 1)
            self.assertGreater(len(blob), 100)
            self.assertEqual(blob[:2], b"PK")
            from openpyxl import load_workbook
            import io

            wb2 = load_workbook(io.BytesIO(blob), read_only=True)
            ws2 = wb2.active
            self.assertEqual(str(ws2.cell(1, 1).value), "材料名称")
            self.assertEqual(str(ws2.cell(2, 1).value), "EXPORT_TEST")
            wb2.close()

            history = list_price_history(limit=5, history_path=root / "history.jsonl")
            self.assertEqual(history[0]["action"], "export_workbook")
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_import_price_workbook(self, _mock_note: object, _mock_reload: object) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_import_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            target_kb = self._make_kb(root)
            import_kb = root / "import_source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价"])
            ws.append(["YKK拉链", "5#", "7元/码"])
            ws.append(["210D涤纶", "152CM", "3元/码"])
            wb.save(import_kb)
            wb.close()

            payload_b64 = base64.b64encode(import_kb.read_bytes()).decode("utf-8")
            out = import_price_kb_workbook(
                filename="import_source.xlsx",
                content_base64=payload_b64,
                kb_path=target_kb,
                history_path=root / "history.jsonl",
                updated_by="tester",
            )
            self.assertTrue(out.get("ok"))
            self.assertEqual(out.get("rows"), 2)

            items, total = list_price_entries(page=1, page_size=20, kb_path=target_kb)
            self.assertEqual(total, 2)
            names = {x["name"] for x in items}
            self.assertIn("YKK拉链", names)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_quote_rows_auto_learns_and_pending_without_overwrite(
        self, _mock_note: object, _mock_reload: object
    ) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            history_path = root / "price_kb_history.jsonl"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            ws.append(["EXISTING_FABRIC", "150CM", "12/Y", "", "active", "seed", "2026-05-19 09:00:00", "seed"])
            wb.save(kb_path)
            wb.close()

            quote = {
                "quote_ready": True,
                "quote_id": "Q-SYNC-001",
                "product_name": "sync test bag",
                "detail_rows": [
                    {"name": "EXISTING_FABRIC", "spec": "150CM", "unit_price": "12/Y"},
                    {"name": "NEW_ACTIVE_WEBBING", "spec": "25MM", "unit_price": "0.5/M"},
                    {"name": "MISSINGBUCKLE", "spec": "777ZZ", "unit_price": "-"},
                    {"name": "EXISTING_FABRIC", "spec": "150CM", "unit_price": "14/Y"},
                ],
            }

            with patch.dict(os.environ, {"PRICE_REVIEW_DATA_DIR": str(root)}):
                summary = sync_quote_detail_rows_to_price_kb(
                    quote,
                    kb_path=kb_path,
                    history_path=history_path,
                    exception_path=root / "price_exceptions.jsonl",
                    updated_by="test_agent",
                )
                from price_kb_paths import quote_sync_suggestions_path

                sugg_path = quote_sync_suggestions_path()
                sugg_text = sugg_path.read_text(encoding="utf-8") if sugg_path.is_file() else ""
            self.assertEqual(summary["auto_inserted"], 1)
            self.assertEqual(summary["pending"], 2)
            self.assertEqual(summary["conflicts"], 1)
            self.assertEqual(summary["skipped"], 1)
            self.assertEqual(_mock_note.call_count, 1)
            self.assertEqual(_mock_reload.call_count, 1)

            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 2)
            existing_rows = [x for x in items if x["name"] == "EXISTING_FABRIC"]
            self.assertEqual(sorted(x["price"] for x in existing_rows), ["12/Y"])
            webbing = [x for x in items if x["name"] == "NEW_ACTIVE_WEBBING"]
            self.assertEqual(len(webbing), 1)
            self.assertEqual(webbing[0]["price"], "0.5/M")

            self.assertNotIn("NEW_ACTIVE_WEBBING", sugg_text)

            pending_rows, pending_total = list_price_exceptions(
                page=1,
                page_size=20,
                exception_path=root / "price_exceptions.jsonl",
            )
            self.assertEqual(pending_total, 2)
            pending_by_name = {x["name"]: x for x in pending_rows}
            self.assertEqual(pending_by_name["MISSINGBUCKLE"]["marker"], AUTO_PENDING_MARKER)
            self.assertEqual(pending_by_name["EXISTING_FABRIC"]["marker"], AUTO_CONFLICT_MARKER)

            kb = get_price_kb(kb_path)
            self.assertIsNone(kb.lookup("MISSINGBUCKLE", "777ZZ", min_score=0.05))
            hit = kb.lookup("NEW_ACTIVE_WEBBING", "25MM", min_score=0.05)
            self.assertIsNotNone(hit)
            self.assertEqual(hit.entry.raw_price, "0.5/M")
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_skips_garbage_symbols_and_name_only_rows(
        self, _mock_note: object, _mock_reload: object
    ) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_guard_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = self._make_kb(root)
            quote = {
                "quote_ready": True,
                "quote_id": "Q-SYNC-GUARD",
                "product_name": "guard test bag",
                "detail_rows": [
                    {"name": "300D??", "spec": "152cm", "unit_price": "10.5?/Y", "usage": "0.2Y"},
                    {"name": "210D????", "spec": "58#", "unit_price": "12.5/?", "usage": "0.1Y"},
                    {"name": "NAME_ONLY_MATERIAL", "spec": "-", "unit_price": "-", "usage": "-"},
                    {"name": "SAFE_WEBBING", "spec": "25MM", "unit_price": "0.5/M", "usage": "1M"},
                ],
            }

            with patch.dict(os.environ, {"PRICE_REVIEW_DATA_DIR": str(root)}):
                summary = sync_quote_detail_rows_to_price_kb(
                    quote,
                    kb_path=kb_path,
                    history_path=root / "history.jsonl",
                    exception_path=root / "price_exceptions.jsonl",
                    updated_by="test_agent",
                )
                from price_kb_paths import quote_sync_suggestions_path

                sugg_path = quote_sync_suggestions_path()
                safe_sugg = sugg_path.read_text(encoding="utf-8") if sugg_path.is_file() else ""

            self.assertEqual(summary["auto_inserted"], 1)
            self.assertEqual(summary["pending"], 0)
            self.assertGreaterEqual(summary["skipped"], 3)
            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 2)
            names = {x["name"] for x in items}
            self.assertIn("SAFE_WEBBING", names)
            self.assertNotIn("SAFE_WEBBING", safe_sugg)
            self.assertNotIn("300D??", names)
            self.assertNotIn("210D????", names)
            self.assertNotIn("NAME_ONLY_MATERIAL", names)
            pending_rows, pending_total = list_price_exceptions(
                page=1,
                page_size=20,
                exception_path=root / "price_exceptions.jsonl",
            )
            self.assertEqual(pending_total, 0)
            self.assertEqual(pending_rows, [])
            self.assertEqual(_mock_note.call_count, 1)
            self.assertEqual(_mock_reload.call_count, 1)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_drops_garbage_quality_rows(
        self, _mock_note: object, _mock_reload: object
    ) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_quality_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = self._make_kb(root)
            with patch.dict(os.environ, {"PRICE_REVIEW_DATA_DIR": str(root)}):
                summary = sync_quote_detail_rows_to_price_kb(
                    {
                        "quote_ready": True,
                        "quote_id": "Q-QUALITY-DROP",
                        "product_name": "quality bag",
                        "detail_rows": [
                            {"name": "侧面的主面", "spec": "-", "unit_price": "1元/个"},
                            {"name": "NEW_ACTIVE_WEBBING", "spec": "25MM", "unit_price": "0.5/M"},
                        ],
                    },
                    kb_path=kb_path,
                    history_path=root / "history.jsonl",
                    exception_path=root / "price_exceptions.jsonl",
                )
            self.assertTrue(summary.get("ignored_test_quote"))
            self.assertEqual(summary.get("suggestions_queued"), 0)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_cleans_role_words_and_skips_usage_sentences(
        self, _mock_note: object, _mock_reload: object
    ) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_roles_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            wb.save(kb_path)
            wb.close()

            with patch.dict(os.environ, {"PRICE_REVIEW_DATA_DIR": str(root)}):
                summary = sync_quote_detail_rows_to_price_kb(
                    {
                        "quote_ready": True,
                        "quote_id": "Q-SYNC-ROLE",
                        "product_name": "role test bag",
                        "detail_rows": [
                            {"name": "210D尼龙（内衬）", "spec": "58#", "unit_price": "12.5/码", "usage": "0.09㎡"},
                            {"name": "国产X-PAC（主体面料）", "spec": "-", "unit_price": "50元/码", "usage": "0.30码"},
                            {"name": "仅用于包体最下方的底部贴片", "spec": "-", "unit_price": "12元/码", "usage": "1套"},
                        ],
                    },
                    kb_path=kb_path,
                    history_path=root / "history.jsonl",
                    exception_path=root / "price_exceptions.jsonl",
                    updated_by="test_agent",
                )
                from price_kb_paths import quote_sync_suggestions_path

                sugg_path = quote_sync_suggestions_path()
                role_sugg = sugg_path.read_text(encoding="utf-8") if sugg_path.is_file() else ""

            self.assertGreaterEqual(summary["auto_inserted"], 2)
            self.assertEqual(summary["pending"], 0)
            self.assertEqual(summary["skipped"], 0)
            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 2)
            names = {x["name"] for x in items}
            self.assertIn("210D尼龙", names)
            self.assertIn("国产X-PAC", names)
            self.assertEqual(_mock_note.call_count, 1)
            self.assertEqual(_mock_reload.call_count, 1)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_checks_full_kb_not_only_first_page(self, _mock_note: object, _mock_reload: object) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_full_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            for idx in range(230):
                ws.append([f"FILLER_{idx:03d}", "A", "1/PCS", "", "active", "", "", "seed"])
            ws.append(["LATE_EXISTING", "TAIL", "9/PCS", "", "active", "after page 200", "", "seed"])
            wb.save(kb_path)
            wb.close()

            summary = sync_quote_detail_rows_to_price_kb(
                {
                    "quote_ready": True,
                    "quote_id": "Q-SYNC-TAIL",
                    "detail_rows": [
                        {"name": "LATE_EXISTING", "spec": "TAIL", "unit_price": "9/PCS"},
                    ],
                },
                kb_path=kb_path,
                history_path=root / "history.jsonl",
                exception_path=root / "price_exceptions.jsonl",
            )
            self.assertEqual(summary.get("suggestions_queued"), 0)
            self.assertEqual(summary["skipped"], 1)
            items, total = list_price_entries(
                page=1,
                page_size=20,
                search_q="LATE_EXISTING",
                kb_path=kb_path,
            )
            self.assertEqual(total, 1)
            self.assertEqual(items[0]["price"], "9/PCS")
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_splits_combined_material_names_before_auto_learn(
        self, _mock_note: object, _mock_reload: object
    ) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_split_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            ws.append(["YKK防水拉链", "-", "1.76", "", "active", "seed", "2026-05-19 09:00:00", "seed"])
            wb.save(kb_path)
            wb.close()

            with patch.dict(os.environ, {"PRICE_REVIEW_DATA_DIR": str(root)}):
                summary = sync_quote_detail_rows_to_price_kb(
                    {
                        "quote_ready": True,
                        "quote_id": "Q-SYNC-SPLIT",
                        "product_name": "sync split bag",
                        "detail_rows": [
                            {"name": "#5尼龙拉链，YKK防水拉链", "spec": "-", "unit_price": "7.5"},
                        ],
                    },
                    kb_path=kb_path,
                    history_path=root / "history.jsonl",
                    exception_path=root / "price_exceptions.jsonl",
                )
            self.assertEqual(summary["pending"], 1)
            self.assertEqual(summary["skipped"], 1)

            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 1)
            names = {x["name"]: x for x in items}
            self.assertNotIn("#5尼龙拉链，YKK防水拉链", names)
            self.assertNotIn("#5尼龙拉链", names)

            pending_rows, pending_total = list_price_exceptions(
                page=1,
                page_size=20,
                exception_path=root / "price_exceptions.jsonl",
            )
            self.assertEqual(pending_total, 1)
            pending_by_name = {x["name"]: x for x in pending_rows}
            self.assertEqual(pending_by_name["#5尼龙拉链"]["status"], "pending")
            self.assertEqual(pending_by_name["#5尼龙拉链"]["marker"], AUTO_PENDING_MARKER)
            self.assertEqual(pending_by_name["#5尼龙拉链"]["price"], "")
            self.assertIn("无法自动分摊", pending_by_name["#5尼龙拉链"]["note"])
            self.assertEqual(_mock_note.call_count, 0)
            self.assertEqual(_mock_reload.call_count, 0)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_approve_price_exception_moves_to_kb(self, _mock_note: object, _mock_reload: object) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_exception_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            history_path = root / "history.jsonl"
            exception_path = root / "price_exceptions.jsonl"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            wb.save(kb_path)
            wb.close()

            with patch.dict(os.environ, {"PRICE_REVIEW_DATA_DIR": str(root)}):
                sync_quote_detail_rows_to_price_kb(
                    {
                        "quote_ready": True,
                        "quote_id": "Q-APPROVE-001",
                        "product_name": "approve test bag",
                        "detail_rows": [
                            {"name": "NEW_PENDING_BUCKLE", "spec": "777ZZ", "unit_price": "-"},
                        ],
                    },
                    kb_path=kb_path,
                    history_path=history_path,
                    exception_path=exception_path,
                )

            pending_rows, pending_total = list_price_exceptions(
                page=1,
                page_size=20,
                exception_path=exception_path,
            )
            self.assertEqual(pending_total, 1)
            exc = pending_rows[0]
            self.assertEqual(exc["name"], "NEW_PENDING_BUCKLE")

            result = approve_price_exception(
                str(exc["exception_id"]),
                {
                    "name": "NEW_PENDING_BUCKLE",
                    "spec": "777ZZ",
                    "price": "2.5/PCS",
                    "status": "active",
                    "updated_by": "admin_tester",
                    "note": "人工确认后入库",
                },
                kb_path=kb_path,
                history_path=history_path,
                exception_path=exception_path,
            )
            self.assertTrue(result.get("ok"))
            self.assertEqual(result["entry"]["status"], "active")
            self.assertEqual(result["entry"]["price"], "2.5/PCS")

            kb_items, kb_total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(kb_total, 1)
            self.assertEqual(kb_items[0]["name"], "NEW_PENDING_BUCKLE")
            self.assertEqual(kb_items[0]["status"], "active")

            resolved_rows, open_total = list_price_exceptions(
                page=1,
                page_size=20,
                status="open",
                exception_path=exception_path,
            )
            self.assertEqual(open_total, 0)
            self.assertEqual(len(resolved_rows), 0)
            self.assertEqual(_mock_note.call_count, 1)
            self.assertEqual(_mock_reload.call_count, 1)
        finally:
            shutil.rmtree(root, ignore_errors=True)

    @patch("price_admin_store.knowledge_reload_hook")
    @patch("price_admin_store.note_kb_disk_write_success")
    def test_sync_skips_estimated_packaging_cost_rows(self, _mock_note: object, _mock_reload: object) -> None:
        root = Path(__file__).resolve().parents[1] / "data" / f"_tmp_price_sync_pack_{uuid.uuid4().hex[:8]}"
        root.mkdir(parents=True, exist_ok=True)
        try:
            kb_path = root / "price_kb.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "材料询价"
            ws.append(["材料名称", "规格大小", "单价", "标记", "状态", "备注", "更新时间", "更新人"])
            ws.append(["EXISTING_FABRIC", "150CM", "12/Y", "", "active", "seed", "2026-05-19 09:00:00", "seed"])
            wb.save(kb_path)
            wb.close()

            summary = sync_quote_detail_rows_to_price_kb(
                {
                    "quote_ready": True,
                    "quote_id": "Q-SYNC-PACK",
                    "product_name": "sync packaging bag",
                    "detail_rows": [
                        {"name": "外纸箱/包装费（系统估算）", "spec": "—", "unit_price": "8.00元/个"},
                        {"name": "纸箱/包装袋（加计）", "spec": "-", "unit_price": "6.00元/个"},
                    ],
                },
                kb_path=kb_path,
                history_path=root / "history.jsonl",
            )
            self.assertEqual(summary.get("suggestions_queued"), 0)
            self.assertEqual(summary["pending"], 0)
            self.assertEqual(summary["conflicts"], 0)
            self.assertEqual(summary["skipped"], 0)
            self.assertEqual(_mock_note.call_count, 0)
            self.assertEqual(_mock_reload.call_count, 0)

            items, total = list_price_entries(page=1, page_size=20, kb_path=kb_path)
            self.assertEqual(total, 1)
            self.assertEqual(items[0]["name"], "EXISTING_FABRIC")
        finally:
            shutil.rmtree(root, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
