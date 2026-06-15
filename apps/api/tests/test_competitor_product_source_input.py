"""competitor_product：Amazon source_input_url 端到端追溯。"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from app.collectors.base import CollectedInfluencer
from app.models.enums import CollectionMode
from app.services.apify_instagram import PostAuthorCandidate
from app.services.competitor_product_discovery import (
    apply_competitor_product_source_context,
    apply_competitor_product_source_to_collected,
    build_candidate_source_meta,
    filter_candidates_by_competitor_caption,
    resolve_amazon_source_input_urls,
)
from app.services.export import build_collection_task_candidates_excel, build_influencer_library_excel
from app.services.platform_types import PlatformCandidateProfile
from app.services.platform_utils import candidate_row_from_profile, profile_to_collected

AMAZON_ORIGINAL = (
    "https://www.amazon.com/Portable-Laundry-Apartment-Drawstring-XL/dp/B0CPF3W9B2/"
    "ref=sr_1_1?tag=test-20"
)
AMAZON_NORMALIZED = "https://www.amazon.com/dp/B0CPF3W9B2"
YOUTUBE_POST = "https://www.youtube.com/watch?v=dQw4w9WgXcQ"


def _competitor_task(**overrides):
    defaults = {
        "collection_mode": CollectionMode.COMPETITOR_PRODUCT.value,
        "input_urls": [AMAZON_NORMALIZED],
        "keywords": ["B0CPF3W9B2"],
        "run_checkpoint": {
            "amazon_product_seeds": [
                {
                    "url": AMAZON_ORIGINAL,
                    "normalized_url": AMAZON_NORMALIZED,
                    "platform": "amazon",
                    "asin": "B0CPF3W9B2",
                    "marketplace": "amazon.com",
                    "source_type": "amazon_product",
                }
            ]
        },
    }
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def test_resolve_amazon_source_input_urls_prefers_normalized_seed():
    task = _competitor_task()
    normalized, original = resolve_amazon_source_input_urls(task)
    assert normalized == AMAZON_NORMALIZED
    assert original == AMAZON_ORIGINAL


@pytest.mark.parametrize("platform", ["youtube", "tiktok", "facebook"])
def test_apply_competitor_source_context_on_non_instagram_profiles(platform: str):
    task = _competitor_task()
    profile = PlatformCandidateProfile(
        platform=platform,
        username="creator",
        profile_url=f"https://example.com/{platform}/creator",
        source_post_url=YOUTUBE_POST if platform == "youtube" else f"https://{platform}.com/video/1",
        source_url=YOUTUBE_POST if platform == "youtube" else f"https://{platform}.com/video/1",
        source_discovery_type="video_channel",
    )
    enriched = apply_competitor_product_source_context(profile, task)
    item = profile_to_collected(enriched)

    assert enriched.source_input_url == AMAZON_NORMALIZED
    assert enriched.source_meta["source_input_url"] == AMAZON_NORMALIZED
    assert enriched.source_meta["amazon_original_url"] == AMAZON_ORIGINAL
    assert item.source_input_url == AMAZON_NORMALIZED
    assert item.source_post_url


def test_candidate_row_and_candidate_export_include_amazon_source_input():
    task = _competitor_task()
    profile = apply_competitor_product_source_context(
        PlatformCandidateProfile(
            platform="youtube",
            username="creator",
            profile_url="https://www.youtube.com/channel/UC123",
            source_post_url=YOUTUBE_POST,
            source_discovery_type="video_channel",
        ),
        task,
    )
    row = candidate_row_from_profile(
        profile,
        status="discovered",
        collection_mode=task.collection_mode,
    )
    assert row["source_post_url"] == YOUTUBE_POST
    assert row["source_input_url"] == AMAZON_NORMALIZED
    assert row["source_meta"]["source_input_url"] == AMAZON_NORMALIZED

    candidate = SimpleNamespace(
        username="creator",
        platform="youtube",
        profile_url=profile.profile_url,
        source_post_url=YOUTUBE_POST,
        source_input_url=AMAZON_NORMALIZED,
        source_meta=row["source_meta"],
        followers_count=1000,
        engagement_rate=2.0,
        status="discovered",
        created_at=datetime(2026, 6, 15, tzinfo=UTC),
    )
    content, _ = build_collection_task_candidates_excel(
        [(candidate, None)],
        task_id=5,
        task_name="Amazon task",
    )
    ws = load_workbook(BytesIO(content)).active
    headers = [cell.value for cell in ws[1]]
    assert ws.cell(row=2, column=headers.index("来源输入链接") + 1).value == AMAZON_NORMALIZED
    assert ws.cell(row=2, column=headers.index("来源作品链接") + 1).value == YOUTUBE_POST


def test_instagram_competitor_candidate_carries_source_input_url():
    task = _competitor_task()
    from app.services.competitor_product_discovery import CompetitorProductInfo, CaptionMatchResult

    info = CompetitorProductInfo(
        asin="B0CPF3W9B2",
        amazon_urls=[AMAZON_NORMALIZED],
    )
    candidate = PostAuthorCandidate(
        username="creator",
        profile_url="https://www.instagram.com/creator/",
        source_post_url="https://www.instagram.com/p/abc/",
        source_caption="Love this amazon find #amazonfinds B0CPF3W9B2",
    )
    matched, _ = filter_candidates_by_competitor_caption(
        [candidate],
        info,
        source_input_url=AMAZON_NORMALIZED,
        amazon_original_url=AMAZON_ORIGINAL,
    )
    assert len(matched) == 1
    assert matched[0].source_input_url == AMAZON_NORMALIZED
    assert matched[0].source_meta["source_input_url"] == AMAZON_NORMALIZED

    meta = build_candidate_source_meta(
        info,
        CaptionMatchResult(matched=True, matched_keywords=["amazon"]),
        source_post_url=candidate.source_post_url,
        source_caption=candidate.source_caption,
        source_input_url=AMAZON_NORMALIZED,
        amazon_original_url=AMAZON_ORIGINAL,
    )
    assert meta["source_input_url"] == AMAZON_NORMALIZED


def test_record_from_collected_resolves_amazon_source_fields():
    from app.services.influencer_source import resolve_source_fields

    post_url, input_url, source_key = resolve_source_fields(
        source_post_url=YOUTUBE_POST,
        source_input_url=AMAZON_NORMALIZED,
        profile_url="https://www.youtube.com/channel/UC123",
    )
    assert post_url == YOUTUBE_POST
    assert input_url == AMAZON_NORMALIZED
    assert source_key


def test_influencer_library_export_includes_amazon_source_input_url():
    influencer = SimpleNamespace(
        id=11,
        username="creator",
        display_name="Creator",
        platform="youtube",
        profile_url="https://www.youtube.com/channel/UC123",
        bio="bio",
        followers_count=1000,
        engagement_rate=2.0,
        final_email="a@example.com",
        email="a@example.com",
        public_email=None,
        business_email=None,
        source_post_url=YOUTUBE_POST,
    )
    sources = {
        11: [
            SimpleNamespace(
                source_post_url=YOUTUBE_POST,
                source_input_url=AMAZON_NORMALIZED,
                task_name="Amazon competitor",
                task_id=5,
                source_platform="youtube",
                collected_at=datetime(2026, 6, 15, tzinfo=UTC),
            )
        ]
    }
    content, _ = build_influencer_library_excel([influencer], sources_by_influencer_id=sources)
    ws = load_workbook(BytesIO(content)).active
    headers = [cell.value for cell in ws[1]]
    assert ws.cell(row=2, column=headers.index("来源输入链接") + 1).value == AMAZON_NORMALIZED
    assert ws.cell(row=2, column=headers.index("来源作品链接") + 1).value == YOUTUBE_POST


def test_apply_competitor_product_source_to_collected_from_task():
    task = _competitor_task()
    item = CollectedInfluencer(
        platform="tiktok",
        username="creator",
        profile_url="https://www.tiktok.com/@creator",
        source_post_url="https://www.tiktok.com/@creator/video/1",
    )
    apply_competitor_product_source_to_collected(item, task)
    assert item.source_input_url == AMAZON_NORMALIZED
