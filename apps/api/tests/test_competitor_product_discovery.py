import asyncio
from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.models.collection_task import CollectionTask
from app.schemas.collection_task import CollectionTaskCreate, CollectionTaskUpdate
from app.services.apify_instagram import PostAuthorCandidate
from app.services.collection_task import CollectionTaskService
from app.services.amazon_url import looks_like_asin
from app.services.competitor_product_discovery import (
    apply_competitor_product_relevance_to_platform_results,
    apply_cross_platform_evidence_to_instagram_item,
    amazon_product_video_candidate_rows,
    build_competitor_search_keywords,
    build_candidate_source_meta,
    build_cross_platform_instagram_probe_queries,
    build_instagram_product_search_queries,
    build_instagram_profile_probe_urls,
    discover_competitor_product_candidates,
    extract_asin_from_text,
    filter_candidates_by_competitor_caption,
    filter_platform_profiles_by_competitor_relevance,
    match_competitor_caption,
    order_competitor_discovery_platforms,
    parse_competitor_product_inputs,
    resolve_competitor_discovery_keywords,
)
from app.services.platform_types import PlatformCandidateProfile
from app.services.platform_types import PlatformDiscoveryResult
from app.collectors.base import CollectedInfluencer
from app.services.export import build_collection_task_candidates_excel
from app.services.collection_funnel import CollectionFunnelStats, build_status_summary
from app.services.competitor_product_discovery import CompetitorProductDiscoveryMeta, CompetitorProductInfo
from app.services.task_candidate import TaskCandidateService
from app.services.task_run_progress import RunCheckpoint
from app.models.enums import CandidateStatus, CollectionMode, CollectionTaskStatus
import pytest


def test_extract_asin_from_amazon_url():
    url = "https://www.amazon.com/dp/B0ABCD1234/ref=sr_1_1"
    assert extract_asin_from_text(url) == "B0ABCD1234"
    assert extract_asin_from_text("B0ABCD1234") == "B0ABCD1234"


def test_parse_competitor_product_inputs_merges_brand_and_amazon():
    task = CollectionTask(
        name="test",
        platform="instagram",
        collection_mode="competitor_product",
        keywords=["brand:Anker", "portable charger"],
        input_urls=["https://www.amazon.com/dp/B0ABCD1234/"],
        category="tech",
    )
    info = parse_competitor_product_inputs(task)
    assert info.asin == "B0ABCD1234"
    assert info.brand == "Anker"
    assert "amazonfinds" in info.search_hashtags
    assert "techfinds" in info.search_hashtags or "gadgetfinds" in info.search_hashtags


def test_amazon_homehive_product_generates_instagram_queries():
    task = CollectionTask(
        name="homehive",
        platform="instagram",
        collection_mode="competitor_product",
        input_urls=["https://www.amazon.com/dp/B0D9W576KQ"],
        category="home organization",
        run_checkpoint={
            "amazon_product_seeds": [
                {
                    "asin": "B0D9W576KQ",
                    "brand": "HOMEHIVE",
                    "product_title": "clear PVC jewelry bags",
                    "strong_keywords": ["clear PVC jewelry bags", "clear jewelry organizer bags"],
                    "broad_category_keywords": ["jewelry storage", "travel jewelry organizer", "organization"],
                    "search_keywords": ["Amazon home organization finds", "Amazon accessories organizer"],
                }
            ]
        },
    )

    queries = build_instagram_product_search_queries(parse_competitor_product_inputs(task))

    assert "HOMEHIVE clear PVC jewelry bags Instagram" in queries
    assert "clear PVC jewelry bags Amazon finds Instagram" in queries
    assert "clear jewelry organizer bags Amazon finds Instagram" in queries
    assert "jewelry storage Amazon finds Instagram" in queries
    assert "travel jewelry organizer Amazon finds Instagram" in queries
    assert "Amazon home organization finds Instagram" in queries
    assert "Amazon accessories organizer Instagram" in queries
    assert "amazonfinds jewelry storage" in queries
    assert "amazonmusthaves organization" in queries
    assert "travelorganizer amazonfinds" in queries


def test_cross_platform_profiles_generate_instagram_probe_queries_and_urls():
    tiktok = PlatformCandidateProfile(
        platform="tiktok",
        username="allstarsteven",
        display_name="All Star Steven",
        profile_url="https://www.tiktok.com/@allstarsteven",
        bio="Amazon finds, link in bio: shopmy.us/allstarsteven steven@example.com",
        source_meta={"source_caption": "HOMEHIVE clear PVC jewelry bags Amazon finds"},
    )
    youtube = PlatformCandidateProfile(
        platform="youtube",
        username="UC123",
        display_name="Sew Simple Home",
        profile_url="https://www.youtube.com/@SewSimpleHome",
        bio="Amazon Associate and sewing blogger",
    )

    queries = build_cross_platform_instagram_probe_queries([tiktok, youtube])
    urls = build_instagram_profile_probe_urls([tiktok, youtube])

    assert "allstarsteven Instagram" in queries
    assert "allstarsteven Amazon finds Instagram" in queries
    assert "allstarsteven link in bio Instagram" in queries
    assert "All Star Steven influencer Instagram" in queries
    assert "Sew Simple Home Instagram" in queries
    assert "Sew Simple Home influencer Instagram" in queries
    assert "https://www.instagram.com/allstarsteven/" in urls
    assert "https://www.instagram.com/sewsimplehome/" in urls


def test_instagram_probe_item_inherits_cross_platform_product_evidence():
    source = PlatformCandidateProfile(
        platform="youtube",
        username="sewsimplehome",
        display_name="Sew Simple Home",
        profile_url="https://www.youtube.com/@SewSimpleHome",
        source_post_url="https://www.youtube.com/watch?v=abc123",
        source_meta={
            "source_caption": "HOMEHIVE clear PVC jewelry bags Amazon finds",
            "amazon_asin": "B0D9W576KQ",
            "matched_keywords": ["clear PVC jewelry bags"],
            "match_reasons": ["same product evidence"],
            "product_match_confidence": "same_product",
        },
    )
    instagram = CollectedInfluencer(
        platform="instagram",
        username="sewsimplehome",
        profile_url="https://www.instagram.com/sewsimplehome/",
        display_name="Sew Simple Home",
    )

    apply_cross_platform_evidence_to_instagram_item(instagram, source)
    meta = getattr(instagram, "source_meta", {}) or {}

    assert instagram.source_discovery_type == "competitor_product_cross_platform_instagram_probe"
    assert meta["evidence_inherited_from_platform"] == "youtube"
    assert meta["evidence_source_profile_url"] == "https://www.youtube.com/@SewSimpleHome"
    assert meta["evidence_source_post_url"] == "https://www.youtube.com/watch?v=abc123"
    assert meta["evidence_source_text"] == "HOMEHIVE clear PVC jewelry bags Amazon finds"
    assert meta["primary_evidence_platform"] == "youtube"
    assert meta["amazon_asin"] == "B0D9W576KQ"


def test_match_competitor_caption_detects_collab_signal():
    info = CompetitorProductInfo(
        asin="B0ABCD1234",
        brand="Anker",
        core_keywords=["charger"],
    )
    match = match_competitor_caption(
        "Love this Anker charger! #amazonfinds #ad link in bio",
        info,
    )
    assert match.matched is True
    assert match.suspected_collab is True
    assert "包含品牌名" in match.match_reasons


def test_filter_candidates_keeps_only_matched_authors():
    info = CompetitorProductInfo(brand="Anker", core_keywords=["charger"])
    candidates = [
        PostAuthorCandidate(
            username="hit",
            profile_url="https://www.instagram.com/hit/",
            source_caption="Anker charger amazonfinds",
            source_post_url="https://www.instagram.com/p/abc/",
        ),
        PostAuthorCandidate(
            username="miss",
            profile_url="https://www.instagram.com/miss/",
            source_caption="random travel photo",
        ),
    ]
    matched, before = filter_candidates_by_competitor_caption(candidates, info)
    assert before == 2
    assert len(matched) == 1
    assert matched[0].username == "hit"
    assert matched[0].source_meta["suspected_collab"] is True


def test_build_status_summary_for_competitor_product():
    info = CompetitorProductInfo(asin="B0ABCD1234", brand="Anker", core_keywords=["charger"])
    competitor_meta = CompetitorProductDiscoveryMeta(
        product_info=info,
        posts_scanned=42,
        authors_matched=7,
    )
    summary = build_status_summary(
        CollectionFunnelStats(discovered_count=7, inserted_count=3, post_count=42),
        status=CollectionTaskStatus.COMPLETED_WITH_RESULTS,
        collection_mode="competitor_product",
        competitor_meta=competitor_meta,
    )
    assert "ASIN B0ABCD1234" in summary
    assert "品牌 Anker" in summary
    assert "42 条帖子" in summary
    assert "7 个疑似推广账号" in summary
    assert "入库 3 个" in summary


def test_collection_task_create_validates_competitor_product():
    task = CollectionTaskCreate(
        name="竞品",
        platform="instagram",
        collection_mode="competitor_product",
        keywords=["portable fan"],
        comment_discovery_enabled=True,
    )
    assert task.comment_discovery_enabled is False

    try:
        CollectionTaskCreate(
            name="竞品",
            platform="instagram",
            collection_mode="competitor_product",
        )
        raise AssertionError("expected validation error")
    except ValueError as exc:
        assert "竞品商品发现" in str(exc)


def test_commerce_only_does_not_match_without_product_signals():
    info = CompetitorProductInfo(brand="Anker", core_keywords=["charger"])
    match = match_competitor_caption("#amazonfinds #ad link in bio sponsored", info)
    assert match.matched is False

    asin_only = CompetitorProductInfo(
        asin="B0ABCD1234",
        amazon_urls=["https://www.amazon.com/dp/B0ABCD1234/"],
        core_keywords=["B0ABCD1234"],
    )
    generic = match_competitor_caption("Love #amazonfinds #ad link in bio", asin_only)
    assert generic.matched is False

    with_asin = match_competitor_caption(
        "Check https://www.amazon.com/dp/B0ABCD1234/ #amazonfinds",
        asin_only,
    )
    assert with_asin.matched is True
    assert with_asin.match_type == "exact_link_match"
    assert with_asin.low_confidence is False


def test_row_from_filtered_and_failed_preserve_source_meta():
    meta = {
        "asin": "B0ABCD1234",
        "brand": "Anker",
        "matched_keywords": ["Anker"],
        "match_reasons": ["包含品牌名"],
        "source_post_url": "https://www.instagram.com/p/abc/",
        "source_caption": "Anker charger",
    }
    filtered = TaskCandidateService.row_from_filtered(
        username="user1",
        profile_url="https://www.instagram.com/user1/",
        failure_reason="below_min_followers",
        source_meta=meta,
    )
    failed = TaskCandidateService.row_from_failed(
        username="user2",
        profile_url="https://www.instagram.com/user2/",
        failure_reason="profile_fetch_failed",
        source_meta=meta,
    )
    assert filtered["source_meta"] == meta
    assert failed["source_meta"] == meta


def test_validate_task_inputs_competitor_product_requires_keywords_or_urls():
    with pytest.raises(ValueError, match="竞品商品发现"):
        CollectionTaskService._validate_task_inputs(
            CollectionMode.COMPETITOR_PRODUCT.value,
            [],
            [],
            False,
            [],
        )
    CollectionTaskService._validate_task_inputs(
        CollectionMode.COMPETITOR_PRODUCT.value,
        ["portable fan"],
        [],
        False,
        [],
    )


def test_competitor_product_update_payload_disables_comment_discovery():
    task = CollectionTask(
        name="竞品",
        platform="instagram",
        collection_mode=CollectionMode.KEYWORD.value,
        keywords=["travel"],
        comment_discovery_enabled=True,
    )
    data = CollectionTaskUpdate(
        collection_mode=CollectionMode.COMPETITOR_PRODUCT,
        keywords=["portable fan"],
        comment_discovery_enabled=True,
    )
    update_data = CollectionTaskService._serialize_task_data(data.model_dump(exclude_unset=True))
    merged_mode = update_data.get("collection_mode", task.collection_mode)
    merged_keywords = update_data.get("keywords", task.keywords or [])
    merged_urls = update_data.get("input_urls", task.input_urls or [])
    CollectionTaskService._validate_task_inputs(
        merged_mode,
        merged_keywords,
        merged_urls,
        update_data.get("email_enabled", task.email_enabled),
        update_data.get("email_recipients", task.email_recipients or []),
    )
    merged_mode_value = (
        merged_mode.value if isinstance(merged_mode, CollectionMode) else merged_mode
    )
    if merged_mode_value == CollectionMode.COMPETITOR_PRODUCT.value:
        update_data["comment_discovery_enabled"] = False
    assert update_data["comment_discovery_enabled"] is False


AMAZON_LAUNDRY_URL = (
    "https://www.amazon.com/Laundry-Washable-Organizer-Drawstring%EF%BC%8CLarge-Essentials/"
    "dp/B0CPF3W9B2/ref=zg_bs_g_3744371_d_sccl_2/138-2111992-2516563?psc=1"
)
AMAZON_HOMEHIVE_URL = "https://www.amazon.com/dp/B0D9W576KQ"
AIR_PURIFIER_TITLE = (
    "Breathing Clean: Top 5 Best Air Purifiers With Washable Filters For A Healthier Home"
)


def _laundry_bag_task_info():
    from app.schemas.collection_task import CollectionTaskCreate

    created = CollectionTaskCreate(
        name="laundry-bag",
        collection_mode="link_import",
        platform="instagram",
        input_urls=[AMAZON_LAUNDRY_URL],
    )
    task = CollectionTask(
        name="laundry-bag",
        platform="multi",
        collection_mode="competitor_product",
        input_urls=created.input_urls,
        keywords=created.keywords,
        run_checkpoint=created.run_checkpoint,
    )
    return parse_competitor_product_inputs(task)


def _homehive_jewelry_bag_task_info():
    from app.schemas.collection_task import CollectionTaskCreate

    created = CollectionTaskCreate(
        name="homehive-jewelry-bags",
        collection_mode="link_import",
        platform="instagram",
        input_urls=[AMAZON_HOMEHIVE_URL],
    )
    task = CollectionTask(
        name="homehive-jewelry-bags",
        platform="multi",
        collection_mode="competitor_product",
        input_urls=created.input_urls,
        keywords=created.keywords,
        run_checkpoint=created.run_checkpoint,
    )
    return parse_competitor_product_inputs(task)


def test_homehive_amazon_link_builds_exact_product_fingerprint():
    info = _homehive_jewelry_bag_task_info()
    search_terms = build_competitor_search_keywords(info)
    assert info.asin == "B0D9W576KQ"
    assert info.brand == "HOMEHIVE"
    assert info.product_category == "jewelry_storage_bag"
    assert "B0D9W576KQ" in search_terms
    assert "HOMEHIVE clear PVC jewelry bags" in search_terms
    assert "HOMEHIVE 20 clear bags" in search_terms
    assert "jewelry organizer" not in search_terms


def test_homehive_search_hashtags_do_not_use_laundry_bag_fallback():
    info = _homehive_jewelry_bag_task_info()
    assert "homehivelaundrybag" not in info.search_hashtags
    assert "homehiveclearpvcjewelrybags" in info.search_hashtags
    assert "homehive20clearbags" in info.search_hashtags


def test_homehive_exact_asin_or_url_is_exact_link_match():
    info = _homehive_jewelry_bag_task_info()
    result = match_competitor_caption(
        "Partnered on these HOMEHIVE jewelry bags https://www.amazon.com/dp/B0D9W576KQ",
        info,
    )
    assert result.matched is True
    assert result.match_type == "exact_link_match"
    assert result.product_match_confidence == "exact"
    assert result.match_score == 100.0


def test_homehive_brand_and_core_title_is_same_product_match():
    info = _homehive_jewelry_bag_task_info()
    result = match_competitor_caption(
        "HOMEHIVE clear PVC jewelry bags keep my rings and earrings dustproof while traveling",
        info,
    )
    assert result.matched is True
    assert result.match_type == "same_product_match"
    assert result.product_match_confidence == "high"
    assert "HOMEHIVE" in result.matched_keywords
    assert "clear PVC jewelry bags" in result.matched_keywords


def test_homehive_variant_attributes_rank_above_same_product():
    info = _homehive_jewelry_bag_task_info()
    same_product = match_competitor_caption(
        "HOMEHIVE clear PVC jewelry bags are my favorite anti tarnish storage",
        info,
    )
    same_variant = match_competitor_caption(
        "HOMEHIVE 20 Clear Bags in clear PVC are perfect zipper jewelry bags",
        info,
    )
    assert same_variant.matched is True
    assert same_variant.match_type == "same_variant_match"
    assert same_variant.match_score > (same_product.match_score or 0)


def test_homehive_weak_category_and_other_brand_are_not_qualified():
    info = _homehive_jewelry_bag_task_info()
    weak = match_competitor_caption(
        "These jewelry organizer travel pouch storage bags are great for necklaces",
        info,
    )
    assert weak.matched is False
    assert weak.match_type == "weak_category_match"
    assert weak.rejected_reason == "weak_category_match"

    other_brand = match_competitor_caption(
        "ACME clear PVC jewelry bags with 20 clear bags for rings and earrings",
        info,
    )
    assert other_brand.matched is False
    assert other_brand.rejected_reason == "missing_brand_for_same_product"


def test_homehive_candidate_meta_explains_same_product_and_filtered_reason():
    info = _homehive_jewelry_bag_task_info()
    candidate = PostAuthorCandidate(
        username="hit",
        profile_url="https://www.instagram.com/hit/",
        source_caption="HOMEHIVE clear PVC jewelry bags for anti tarnish earring storage",
        source_post_url="https://www.instagram.com/p/hit/",
    )
    matched, before = filter_candidates_by_competitor_caption([candidate], info)
    assert before == 1
    assert len(matched) == 1
    meta = matched[0].source_meta
    assert meta["amazon_asin"] == "B0D9W576KQ"
    assert meta["amazon_brand"] == "HOMEHIVE"
    assert meta["amazon_product_title"]
    assert meta["match_type"] == "same_product_match"
    assert "clear PVC jewelry bags" in meta["matched_phrases"]
    assert meta["missing_required_phrases"] == []
    assert "HOMEHIVE" in meta["selected_reason"]

    export_candidate = SimpleNamespace(
        username="hit",
        platform="instagram",
        profile_url="https://www.instagram.com/hit/",
        source_post_url="https://www.instagram.com/p/hit/",
        source_input_url=AMAZON_HOMEHIVE_URL,
        source_meta=meta,
        followers_count=1000,
        engagement_rate=2.0,
        status="discovered",
        created_at=datetime(2026, 6, 15, tzinfo=UTC),
    )
    content, _ = build_collection_task_candidates_excel(
        [(export_candidate, None)],
        task_id=7,
        task_name="HOMEHIVE",
    )
    ws = load_workbook(BytesIO(content)).active
    headers = [cell.value for cell in ws[1]]
    exported_reason = ws.cell(row=2, column=headers.index("商品匹配说明") + 1).value
    assert exported_reason == meta["selected_reason"]

    rejected = match_competitor_caption("jewelry organizer travel pouch", info)
    rejected_meta = build_candidate_source_meta(
        info,
        rejected,
        source_post_url="https://www.instagram.com/p/miss/",
        source_caption="jewelry organizer travel pouch",
    )
    assert rejected_meta["match_type"] == "weak_category_match"
    assert rejected_meta["missing_required_phrases"]
    assert "仅命中" in rejected_meta["selected_reason"]


def test_amazon_homehive_instagram_keyword_fallback_progress_proxy(monkeypatch):
    import app.services.keyword_discovery as keyword_discovery

    created = CollectionTaskCreate(
        name="homehive",
        collection_mode="link_import",
        platform="instagram",
        input_urls=[AMAZON_HOMEHIVE_URL],
    )
    task = CollectionTask(
        name="homehive",
        platform="instagram",
        collection_mode=CollectionMode.COMPETITOR_PRODUCT.value,
        input_urls=created.input_urls,
        keywords=created.keywords,
        run_checkpoint=created.run_checkpoint,
    )

    class FakeDB:
        async def commit(self):
            return None

    async def fake_discover_post_authors_from_hashtags(tags, limit=100):
        return SimpleNamespace(
            candidates=[],
            errors=[],
            post_count=0,
            post_urls=[],
        )

    async def fail_if_keyword_discovery_commits_progress(db, progress_task, **kwargs):
        raise AssertionError("competitor product Instagram keyword fallback should not commit progress")

    monkeypatch.setattr(keyword_discovery, "ensure_instagram_provider_ready", lambda: None)
    monkeypatch.setattr(
        keyword_discovery,
        "discover_post_authors_from_hashtags",
        fake_discover_post_authors_from_hashtags,
    )
    monkeypatch.setattr(keyword_discovery, "update_task_progress", fail_if_keyword_discovery_commits_progress)

    checkpoint = RunCheckpoint()
    result = asyncio.run(
        discover_competitor_product_candidates(
            task,
            checkpoint=checkpoint,
            db=FakeDB(),
        )
    )

    assert result.errors == []
    fallback = task.run_checkpoint["competitor_product_instagram_fallback"]
    assert "HOMEHIVE clear PVC jewelry bags Instagram" in fallback["product_instagram_queries"]
    checkpoint_fallback = checkpoint.to_dict()["competitor_product_instagram_fallback"]
    assert "HOMEHIVE clear PVC jewelry bags Instagram" in checkpoint_fallback["product_instagram_queries"]


def test_laundry_bag_positive_and_negative_caption_relevance():
    info = _laundry_bag_task_info()
    for caption in (
        "Amazon travel laundry bag review",
        "drawstring laundry bag for college dorm",
        "dirty clothes organizer laundry hamper bag",
        "Aegero laundry bag",
        "#amazonfinds laundry bag",
    ):
        assert match_competitor_caption(caption, info).matched is True
    for caption in (
        AIR_PURIFIER_TITLE,
        "Best nut milk makers 2026",
        "Installing a lake-based geothermal heating system",
        "Best washable filters for air purifier",
        "Amazon home essentials you need",
        "This washable organizer is great",
        "Best filter machine for home",
        "Best replacement filter products",
        "Best appliance deals 2026",
        "Amazon appliance organizer",
        "Laundry detergent organizer",
        "Washing machine essentials",
        "Laundry room organizer shelf",
        "Just doing laundry today",
        "Love Aegero products",
    ):
        result = match_competitor_caption(caption, info)
        assert result.matched is False
        assert result.rejected_reason


def test_laundry_bag_search_keywords_exclude_broad_single_tokens():
    info = _laundry_bag_task_info()
    broad = {"laundry", "washable", "organizer", "essentials", "large", "bag", "home", "products", "finds"}
    search_terms = build_competitor_search_keywords(info)
    assert len(search_terms) <= 8
    for term in search_terms:
        if looks_like_asin(term):
            continue
        assert " " in term
        assert term.lower() not in broad
    assert "laundry bag" in search_terms or "travel laundry bag" in search_terms


def test_laundry_bag_amazon_product_video_creator_is_strong_source_candidate():
    created = CollectionTaskCreate(
        name="aegero",
        collection_mode="link_import",
        platform="instagram",
        input_urls=[AMAZON_LAUNDRY_URL],
    )
    task = CollectionTask(
        name="aegero",
        platform="multi",
        platforms=["youtube", "tiktok", "instagram"],
        collection_mode=CollectionMode.COMPETITOR_PRODUCT.value,
        input_urls=created.input_urls,
        keywords=created.keywords,
        run_checkpoint={
            **created.run_checkpoint,
            "amazon_product_page_videos": [
                {
                    "asin": "B0CPF3W9B2",
                    "creator_name": "Kuzzin Vinny Reviews",
                    "video_title": "Aegero 2 Pack XL Travel Laundry Bag Review",
                    "video_source": "customer_review_video",
                    "review_url": "https://www.amazon.com/review/R123",
                    "video_url": "https://www.amazon.com/vdp/abc123",
                    "text": "Aegero travel laundry bag B0CPF3W9B2",
                }
            ],
        },
    )

    rows = amazon_product_video_candidate_rows(task)

    assert len(rows) == 1
    row = rows[0]
    assert row["username"] == "Kuzzin Vinny Reviews"
    assert row["platform"] == "amazon"
    assert row["status"] == CandidateStatus.PENDING_PROFILE.value
    assert row["failure_reason"] == "amazon_product_page_strong_lead"
    assert row["source_post_url"] == "https://www.amazon.com/vdp/abc123"
    assert row["source_input_url"] == "https://www.amazon.com/dp/B0CPF3W9B2"
    assert row["source_meta"]["amazon_asin"] == "B0CPF3W9B2"
    assert row["source_meta"]["amazon_creator_name"] == "Kuzzin Vinny Reviews"
    assert row["source_meta"]["video_source"] == "customer_review_video"
    assert row["source_meta"]["product_match_confidence"] == "exact"
    assert row["source_meta"]["match_type"] == "exact_link_match"
    assert "Amazon 商品页强线索" in row["failure_detail"]


def test_laundry_room_organizer_passes_when_laundry_bag_present():
    info = _laundry_bag_task_info()
    result = match_competitor_caption(
        "Laundry room organizer shelf plus my favorite travel laundry bag",
        info,
    )
    assert result.matched is True


def test_single_laundry_and_brand_only_rejected():
    info = _laundry_bag_task_info()
    laundry_only = match_competitor_caption("laundry", info)
    assert laundry_only.matched is False
    brand_only = match_competitor_caption("Check out Aegero", info)
    assert brand_only.matched is False
    assert brand_only.rejected_reason == "brand_only_no_category"


def test_competitor_discovery_keywords_limited_and_platform_order():
    task = CollectionTask(
        name="laundry-bag",
        platform="multi",
        collection_mode="competitor_product",
        input_urls=[AMAZON_LAUNDRY_URL],
        keywords=[],
    )
    keywords = resolve_competitor_discovery_keywords(task)
    assert len(keywords) <= 8
    assert "laundry bag" in keywords
    ordered = order_competitor_discovery_platforms(["youtube", "tiktok", "facebook"])
    assert ordered.index("tiktok") < ordered.index("youtube")


def test_youtube_air_purifier_profile_not_kept_for_laundry_bag():
    info = _laundry_bag_task_info()
    bad = PlatformCandidateProfile(
        platform="youtube",
        username="air_purifier",
        profile_url="https://www.youtube.com/channel/bad",
        recent_post_titles=[AIR_PURIFIER_TITLE],
        source_post_url="https://www.youtube.com/watch?v=rlu7RuLeb5I",
    )
    good = PlatformCandidateProfile(
        platform="youtube",
        username="laundry_creator",
        profile_url="https://www.youtube.com/channel/good",
        recent_post_titles=["Amazon travel laundry bag review"],
        source_post_url="https://www.youtube.com/watch?v=good123",
    )
    kept, _ = filter_platform_profiles_by_competitor_relevance([bad, good], info)
    assert [p.username for p in kept] == ["laundry_creator"]


def test_platform_same_product_filter_records_rejected_candidates_for_diagnosis():
    from app.services.platform_utils import profile_to_collected

    created = CollectionTaskCreate(
        name="homehive",
        collection_mode="link_import",
        platform="instagram",
        input_urls=[AMAZON_HOMEHIVE_URL],
    )
    task = CollectionTask(
        name="homehive",
        platform="multi",
        platforms=["youtube"],
        collection_mode=CollectionMode.COMPETITOR_PRODUCT.value,
        input_urls=created.input_urls,
        keywords=created.keywords,
        run_checkpoint=created.run_checkpoint,
    )
    profile = PlatformCandidateProfile(
        platform="youtube",
        username="other_brand",
        profile_url="https://www.youtube.com/@other-brand",
        followers_count=45_000,
        engagement_rate=2.5,
        recent_post_titles=["ACME clear PVC jewelry bags with 20 clear bags"],
        source_post_url="https://www.youtube.com/watch?v=other",
        source_meta={"source_keyword": "HOMEHIVE clear PVC jewelry bags"},
    )
    result = PlatformDiscoveryResult(
        platform="youtube",
        profiles=[profile],
        items=[profile_to_collected(profile)],
        discovered_count=1,
        deduped_count=1,
        profile_fetched_count=1,
        api_requests=1,
    )

    apply_competitor_product_relevance_to_platform_results([result], task)

    assert result.items == []
    assert result.profiles == []
    assert result.discovered_count == 1
    assert result.deduped_count == 1
    assert result.candidate_rows
    row = result.candidate_rows[0]
    assert row["status"] == CandidateStatus.FILTERED_OUT.value
    assert row["failure_reason"] == "no_same_product_match"
    assert row["platform"] == "youtube"
    assert row["source_input_url"] == AMAZON_HOMEHIVE_URL
    assert row["source_meta"]["amazon_asin"] == "B0D9W576KQ"
    assert row["source_meta"]["rejected_reason"] == "missing_brand_for_same_product"
