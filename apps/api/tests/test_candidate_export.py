"""候选池 Excel 导出单元测试。"""

from datetime import UTC, datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

import app.main  # noqa: F401

from app.services.export import (
    CANDIDATE_BUSINESS_EXPORT_COLUMNS,
    INFLUENCER_LIBRARY_EXPORT_COLUMNS,
    build_collection_task_candidates_excel,
    build_influencer_excel,
    build_influencer_library_excel,
)

REMOVED_HEADERS = {
    "来源类型",
    "发现方式",
    "状态",
    "未入库原因",
    "来源红人 ID",
    "红人 ID",
    "创建时间",
    "原因详情",
    "来源评论",
    "最终邮箱",
}


def _candidate(**kwargs):
    defaults = {
        "username": "travel_creator",
        "platform": "instagram",
        "profile_url": "https://www.instagram.com/travel_creator/",
        "source_type": "hashtag_post_author",
        "source_discovery_type": "post_author",
        "followers_count": 50_000,
        "engagement_rate": 2.5,
        "status": "inserted",
        "failure_reason": None,
        "failure_detail": None,
        "source_post_url": "https://www.instagram.com/p/abc/",
        "source_comment_text": None,
        "influencer_id": 99,
        "created_at": datetime(2026, 6, 4, 12, 0, 0, tzinfo=UTC),
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _influencer(**kwargs):
    defaults = {
        "username": "travel_creator",
        "platform": "instagram",
        "display_name": "Travel Creator",
        "profile_url": "https://www.instagram.com/travel_creator/",
        "bio": "Lifestyle creator",
        "email": "a@example.com",
        "final_email": "a@example.com",
        "business_email": "biz@example.com",
        "public_email": "public@example.com",
        "website": "https://example.com",
        "linktree_url": "https://linktr.ee/travel",
        "whatsapp": "+123",
        "telegram": "@travel",
        "other_social_links": [
            {
                "type": "amazon_storefront",
                "label": "Amazon storefront",
                "url": "https://amzn.to/3XENIP0",
            }
        ],
        "country": "US",
        "language": "en",
        "category": "Travel",
        "niche": "Luxury travel",
        "followers_count": 50_000,
        "avg_views": 12_000,
        "avg_likes": 800,
        "avg_comments": 45,
        "engagement_rate": 2.5,
        "score": 85.0,
        "final_priority": "A",
        "product_fit": 88.0,
        "engagement_score": 80.0,
        "content_match_score": 75.0,
        "commercial_signal_score": 70.0,
        "contactability_score": 65.0,
        "risk_score": 18.0,
        "risk_level": "low",
        "estimated_collab_price": "$500-800",
        "roi_forecast": 2.4,
        "collaboration_formats": ["Reels", "Story"],
        "content_topics": ["travel", "hotel"],
        "tags": ["luxury"],
        "ai_summary": "适合高端酒店合作",
        "ai_collaboration_suggestion": "建议做 Reels 开箱",
        "ai_outreach_message": "Hi, love your travel content",
        "score_reason": "高互动且内容匹配",
        "source_post_url": "https://www.instagram.com/p/influencer-source/",
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _load_sheet(content: bytes):
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    return ws, headers


def test_build_collection_task_candidates_excel_filename_and_business_headers():
    candidate = _candidate()
    influencer = _influencer()
    content, filename = build_collection_task_candidates_excel(
        [(candidate, influencer)],
        task_id=42,
    )
    assert filename == "collection-task-42-candidates.xlsx"
    assert len(content) > 500

    ws, headers = _load_sheet(content)
    expected_headers = [label for _, label, _ in CANDIDATE_BUSINESS_EXPORT_COLUMNS]
    assert headers == expected_headers
    assert not REMOVED_HEADERS.intersection(headers)
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == f"A1:{ws.cell(row=1, column=len(headers)).column_letter}2"


def test_build_collection_task_candidates_excel_influencer_fields_and_formats():
    candidate = _candidate()
    influencer = _influencer()
    content, _ = build_collection_task_candidates_excel([(candidate, influencer)], task_id=42)
    ws, headers = _load_sheet(content)

    profile_col = headers.index("主页链接") + 1
    followers_col = headers.index("粉丝数") + 1
    engagement_col = headers.index("互动率") + 1
    bio_col = headers.index("简介") + 1
    source_col = headers.index("来源作品链接") + 1

    profile_cell = ws.cell(row=2, column=profile_col)
    assert profile_cell.value == influencer.profile_url
    assert profile_cell.hyperlink.target == influencer.profile_url

    followers_cell = ws.cell(row=2, column=followers_col)
    assert followers_cell.value == 50_000
    assert followers_cell.number_format == "0"
    assert isinstance(followers_cell.value, int)

    engagement_cell = ws.cell(row=2, column=engagement_col)
    assert engagement_cell.value == 2.5
    assert engagement_cell.number_format == "0.00"

    bio_cell = ws.cell(row=2, column=bio_col)
    assert bio_cell.value == influencer.bio
    assert bio_cell.alignment.wrap_text is True

    source_cell = ws.cell(row=2, column=source_col)
    assert source_cell.value == candidate.source_post_url
    assert source_cell.hyperlink.target == candidate.source_post_url

    assert ws.cell(row=2, column=headers.index("AI 推荐理由") + 1).value == influencer.ai_summary
    assert ws.cell(row=2, column=headers.index("内容标签") + 1).value == "travel, hotel"
    assert ws.cell(row=2, column=headers.index("合作形式") + 1).value == "Reels, Story"
    assert ws.cell(row=2, column=headers.index("其他外链") + 1).value == "Amazon storefront: https://amzn.to/3XENIP0"
    assert ws.cell(row=2, column=headers.index("联系方式摘要") + 1).value == influencer.final_email
    assert ws.cell(row=2, column=headers.index("价值分层") + 1).value == "可直接外联"
    assert ws.cell(row=2, column=headers.index("推荐动作") + 1).value == "直接联系"


def test_build_collection_task_candidates_excel_without_influencer():
    candidate = _candidate(
        influencer_id=None,
        status="filtered_out",
        failure_reason="below_min_followers",
        followers_count=12_345,
        engagement_rate=1.2,
    )
    content, filename = build_collection_task_candidates_excel([(candidate, None)], task_id=7)
    assert filename == "collection-task-7-candidates.xlsx"

    ws, headers = _load_sheet(content)
    assert ws.cell(row=2, column=1).value == candidate.username
    assert ws.cell(row=2, column=headers.index("主页链接") + 1).value == candidate.profile_url
    assert ws.cell(row=2, column=headers.index("粉丝数") + 1).value == 12_345
    assert ws.cell(row=2, column=headers.index("互动率") + 1).value == 1.2
    assert ws.cell(row=2, column=headers.index("来源作品链接") + 1).value == candidate.source_post_url
    assert ws.cell(row=2, column=headers.index("昵称") + 1).value is None
    assert ws.cell(row=2, column=headers.index("邮箱") + 1).value is None


def test_build_collection_task_candidates_excel_youtube_platform_header():
    candidate = _candidate(
        username="UCAPrhJwVweWZASGEPoC1Sdw",
        platform="youtube",
        profile_url="https://www.youtube.com/channel/UCAPrhJwVweWZASGEPoC1Sdw",
        source_post_url=None,
    )
    influencer = _influencer(
        username="UCAPrhJwVweWZASGEPoC1Sdw",
        platform="youtube",
        display_name="Pleasant Green",
        profile_url="https://www.youtube.com/channel/UCAPrhJwVweWZASGEPoC1Sdw",
    )
    content, _ = build_collection_task_candidates_excel([(candidate, influencer)], task_id=1)
    ws, headers = _load_sheet(content)

    assert "Instagram 主页" not in headers
    assert headers.index("平台") + 1
    assert ws.cell(row=2, column=headers.index("平台") + 1).value == "youtube"
    profile_cell = ws.cell(row=2, column=headers.index("主页链接") + 1)
    assert profile_cell.value == influencer.profile_url
    assert profile_cell.hyperlink.target == influencer.profile_url


def test_build_influencer_excel_formats_other_social_links():
    influencer = _influencer(
        platform="youtube",
        username="UCsommer",
        display_name="The Sommer Home",
        profile_url="https://www.youtube.com/channel/UCsommer",
        linktree_url="https://lnktr.ee/TheSommerHomeYT",
        other_social_links=[{"type": "linktree", "label": "Shop", "url": "https://lnktr.ee/TheSommerHomeYT"}],
    )
    content, _ = build_influencer_excel([influencer])
    ws, headers = _load_sheet(content)
    linktree_col = headers.index("Linktree/链接页") + 1
    other_col = headers.index("其他外链") + 1
    assert ws.cell(row=2, column=linktree_col).value == "https://lnktr.ee/TheSommerHomeYT"
    assert ws.cell(row=2, column=other_col).value == "Shop: https://lnktr.ee/TheSommerHomeYT"


def test_build_influencer_library_excel_uses_source_map():
    from app.services.export import build_influencer_library_excel

    influencer = SimpleNamespace(
        id=7,
        username="creator_a",
        display_name="Creator A",
        platform="tiktok",
        profile_url="https://www.tiktok.com/@creator_a",
        bio="bio",
        followers_count=1000,
        engagement_rate=3.2,
        final_email="a@example.com",
        email="a@example.com",
        public_email=None,
        business_email=None,
        source_post_url=None,
    )
    sources = {
        7: [
            SimpleNamespace(
                source_post_url="https://www.tiktok.com/@creator_a/video/111",
                source_input_url="https://vm.tiktok.com/111",
                task_name="导入任务",
                task_id=3,
                source_platform="tiktok",
                collected_at=datetime(2026, 6, 4, 12, 0, 0, tzinfo=UTC),
            )
        ]
    }
    content, _ = build_influencer_library_excel([influencer], sources_by_influencer_id=sources)
    ws, headers = _load_sheet(content)
    assert ws.cell(row=2, column=headers.index("来源作品链接") + 1).value.endswith("/video/111")
    assert "vm.tiktok.com/111" in ws.cell(row=2, column=headers.index("来源输入链接") + 1).value
    assert ws.cell(row=2, column=headers.index("来源任务") + 1).value == "导入任务"


def test_build_influencer_library_excel_includes_source_columns():
    influencer = _influencer()
    content, _ = build_influencer_library_excel([influencer])
    ws, headers = _load_sheet(content)

    expected_headers = [label for _, label, _ in INFLUENCER_LIBRARY_EXPORT_COLUMNS]
    assert headers == expected_headers
    assert "来源作品链接" in headers
    assert "来源输入链接" in headers
    assert "来源任务" in headers
    assert "采集时间" in headers

    profile_cell = ws.cell(row=2, column=headers.index("主页链接") + 1)
    assert profile_cell.value == influencer.profile_url
    assert profile_cell.hyperlink.target == influencer.profile_url
    assert ws.cell(row=2, column=headers.index("邮箱") + 1).value == influencer.final_email


def test_build_influencer_library_excel_email_fallback_order():
    influencer = _influencer(
        final_email=None,
        email="primary@example.com",
        public_email="public@example.com",
        business_email="biz@example.com",
    )
    content, _ = build_influencer_library_excel([influencer])
    ws, headers = _load_sheet(content)
    assert ws.cell(row=2, column=headers.index("邮箱") + 1).value == "primary@example.com"

    influencer_public_only = _influencer(
        final_email=None,
        email=None,
        public_email="public@example.com",
        business_email="biz@example.com",
    )
    content_public, _ = build_influencer_library_excel([influencer_public_only])
    ws_public, headers_public = _load_sheet(content_public)
    assert ws_public.cell(row=2, column=headers_public.index("邮箱") + 1).value == "public@example.com"


def test_ltk_seed_enrichment_candidate_export_shows_final_platform_and_ltk_source():
    ltk_url = "https://www.shopltk.com/explore/apieceofmyhaven?utm_source=ig"
    enrichment_meta = {
        "link_seed_platform": "ltk",
        "primary_platform": "instagram",
        "enrichment_attempted": True,
        "is_valuable": True,
    }
    candidate = _candidate(
        username="apieceofmyhaven",
        platform="instagram",
        profile_url="https://www.instagram.com/apieceofmyhaven/",
        source_input_url=ltk_url,
        source_post_url=None,
        source_meta={
            "source_input_url": ltk_url,
            "link_seed_enrichment": enrichment_meta,
        },
        source_type="input_profile",
        source_discovery_type="url_profile",
    )
    influencer = _influencer(
        username="apieceofmyhaven",
        platform="instagram",
        profile_url="https://www.instagram.com/apieceofmyhaven/",
    )
    content, _ = build_collection_task_candidates_excel([(candidate, influencer)], task_id=99)
    ws, headers = _load_sheet(content)

    assert ws.cell(row=2, column=headers.index("平台") + 1).value == "Instagram"
    profile_cell = ws.cell(row=2, column=headers.index("主页链接") + 1)
    assert profile_cell.value == influencer.profile_url
    assert ws.cell(row=2, column=headers.index("来源平台") + 1).value == "LTK"
    source_input_cell = ws.cell(row=2, column=headers.index("来源输入链接") + 1)
    assert source_input_cell.value == ltk_url
    assert "utm_source=ig" in source_input_cell.value
    assert ws.cell(row=2, column=headers.index("Seed 平台") + 1).value == "LTK"
    status = ws.cell(row=2, column=headers.index("Seed 补全状态") + 1).value
    assert "LTK seed 补全为 Instagram" in status


def test_influencer_library_ltk_seed_export_shows_source_and_final_platform():
    ltk_url = "https://www.shopltk.com/explore/apieceofmyhaven"
    influencer = SimpleNamespace(
        id=8,
        username="apieceofmyhaven",
        display_name="Haven",
        platform="instagram",
        profile_url="https://www.instagram.com/apieceofmyhaven/",
        bio="bio",
        followers_count=50_000,
        engagement_rate=2.1,
        final_email="a@example.com",
        email="a@example.com",
        public_email=None,
        business_email=None,
        source_post_url=None,
    )
    sources = {
        8: [
            SimpleNamespace(
                source_post_url=None,
                source_input_url=ltk_url,
                task_name="ltk-import",
                task_id=10,
                source_platform="ltk",
                collected_at=datetime(2026, 6, 4, 12, 0, 0, tzinfo=UTC),
            )
        ]
    }
    content, _ = build_influencer_library_excel([influencer], sources_by_influencer_id=sources)
    ws, headers = _load_sheet(content)

    assert ws.cell(row=2, column=headers.index("平台") + 1).value == "Instagram"
    assert ws.cell(row=2, column=headers.index("来源平台") + 1).value == "LTK"
    assert ws.cell(row=2, column=headers.index("来源输入链接") + 1).value == ltk_url
    assert ws.cell(row=2, column=headers.index("Seed 平台") + 1).value == "LTK"
    assert "LTK seed 补全为 Instagram" in ws.cell(row=2, column=headers.index("Seed 补全状态") + 1).value
