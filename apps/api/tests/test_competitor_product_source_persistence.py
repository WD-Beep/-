"""competitor_product source_input_url：真实数据库写入与 API 回归。"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime
from io import BytesIO

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.db.session import async_session_factory
from app.main import app
from app.models.collection_task import CollectionTask
from app.models.collection_task_candidate import CollectionTaskCandidate
from app.models.enums import CandidateStatus, CollectionMode, CollectionTaskStatus
from app.models.global_influencer_profile import GlobalInfluencerProfile
from app.models.product_influencer import ProductInfluencer
from app.schemas.collection_task import collection_task_candidate_read
from app.services.competitor_product_discovery import apply_competitor_product_source_context
from app.services.export import build_collection_task_candidates_excel, build_influencer_library_excel
from app.services.influencer_source import InfluencerSourceService
from app.services.platform_types import PlatformCandidateProfile
from app.services.platform_utils import candidate_row_from_profile
from app.services.task_candidate import TaskCandidateService

AMAZON_ORIGINAL = (
    "https://www.amazon.com/Laundry-Washable-Organizer-Drawstring%EF%BC%8CLarge-Essentials/"
    "dp/B0CPF3W9B2/ref=zg_bs_g_3744371_d_sccl_2/138-2111992-2516563?psc=1"
)
AMAZON_NORMALIZED = "https://www.amazon.com/dp/B0CPF3W9B2"
YOUTUBE_POST = "https://www.youtube.com/watch?v=dQw4w9WgXcQ"


def _competitor_task(**kwargs) -> CollectionTask:
    suffix = uuid.uuid4().hex[:8]
    defaults = {
        "name": f"amazon-competitor-{suffix}",
        "product_id": 1,
        "collection_mode": CollectionMode.COMPETITOR_PRODUCT.value,
        "platform": "multi",
        "platforms": ["instagram", "youtube", "tiktok", "facebook"],
        "keywords": ["B0CPF3W9B2"],
        "input_urls": [AMAZON_NORMALIZED],
        "status": CollectionTaskStatus.COMPLETED_NO_RESULTS.value,
        "last_run_at": datetime.now(UTC),
        "inserted_count": 0,
        "result_count": 0,
        "discovered_count": 1,
        "profile_fetched_count": 0,
        "is_archived": False,
        "run_checkpoint": {
            "amazon_product_seeds": [
                {
                    "url": AMAZON_ORIGINAL,
                    "normalized_url": AMAZON_NORMALIZED,
                    "platform": "amazon",
                    "asin": "B0CPF3W9B2",
                    "marketplace": "amazon.com",
                    "source_type": "amazon_product",
                    "product_keywords": ["laundry", "organizer", "drawstring"],
                }
            ]
        },
    }
    defaults.update(kwargs)
    return CollectionTask(**defaults)


def _youtube_candidate_row(task: CollectionTask) -> dict:
    profile = apply_competitor_product_source_context(
        PlatformCandidateProfile(
            platform="youtube",
            username="yt_creator",
            profile_url="https://www.youtube.com/channel/UC1234567890",
            source_post_url=YOUTUBE_POST,
            source_url=YOUTUBE_POST,
            source_discovery_type="video_channel",
        ),
        task,
    )
    return candidate_row_from_profile(
        profile,
        status=CandidateStatus.DISCOVERED.value,
        collection_mode=task.collection_mode,
    )


@pytest.mark.anyio
async def test_bulk_insert_persists_source_input_url_on_candidate_orm():
    async with async_session_factory() as db:
        task = _competitor_task()
        db.add(task)
        await db.flush()
        run_at = datetime.now(UTC)
        row = _youtube_candidate_row(task)
        await TaskCandidateService.bulk_insert(
            db,
            task.id,
            [row],
            run_at=run_at,
            product_id=1,
        )
        await db.flush()
        candidate = (
            await db.execute(
                select(CollectionTaskCandidate).where(CollectionTaskCandidate.task_id == task.id)
            )
        ).scalar_one()
        assert candidate.source_post_url == YOUTUBE_POST
        assert candidate.source_input_url == AMAZON_NORMALIZED
        read = collection_task_candidate_read(candidate)
        assert read.source_input_url == AMAZON_NORMALIZED
        assert read.source_post_url == YOUTUBE_POST
        await db.rollback()


@pytest.mark.anyio
async def test_candidates_api_returns_source_input_url_from_db():
    task_id: int
    async with async_session_factory() as db:
        task = _competitor_task()
        db.add(task)
        await db.flush()
        task_id = task.id
        await TaskCandidateService.bulk_insert(
            db,
            task.id,
            [_youtube_candidate_row(task)],
            run_at=datetime.now(UTC),
            product_id=1,
        )
        await db.commit()

    transport = ASGITransport(app=app)
    try:
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            response = await client.get(
                f"/api/collection-tasks/{task_id}/candidates",
                headers={"X-User-Id": "1", "X-Product-Id": "1"},
            )
            assert response.status_code == 200, response.text
            item = response.json()["items"][0]
            assert item["source_post_url"] == YOUTUBE_POST
            assert item["source_input_url"] == AMAZON_NORMALIZED
    finally:
        async with async_session_factory() as db:
            row = await db.get(CollectionTask, task_id)
            if row:
                await db.delete(row)
                await db.commit()


@pytest.mark.anyio
async def test_product_influencer_source_persists_amazon_source_input_url():
    from app.collectors.base import CollectedInfluencer

    suffix = uuid.uuid4().hex[:8]
    async with async_session_factory() as db:
        task = _competitor_task()
        db.add(task)
        await db.flush()
        global_row = GlobalInfluencerProfile(
            platform="youtube",
            username=f"yt_{suffix}",
            normalized_username=f"yt_{suffix}",
            profile_url=f"https://www.youtube.com/channel/UC{suffix}",
            normalized_profile_url=f"https://www.youtube.com/channel/UC{suffix}",
        )
        db.add(global_row)
        await db.flush()
        product_row = ProductInfluencer(
            product_id=1,
            global_influencer_id=global_row.id,
        )
        db.add(product_row)
        await db.flush()
        item = CollectedInfluencer(
            platform="youtube",
            username=f"yt_{suffix}",
            profile_url=global_row.profile_url,
            source_post_url=YOUTUBE_POST,
            source_input_url=AMAZON_NORMALIZED,
        )
        source = await InfluencerSourceService.record_from_collected(
            db,
            product_row,
            item,
            task=task,
            run_at=datetime.now(UTC),
        )
        await db.flush()
        assert source is not None
        assert source.source_post_url == YOUTUBE_POST
        assert source.source_input_url == AMAZON_NORMALIZED
        await db.rollback()


@pytest.mark.anyio
async def test_candidate_excel_export_reads_persisted_source_input_url():
    async with async_session_factory() as db:
        task = _competitor_task()
        db.add(task)
        await db.flush()
        await TaskCandidateService.bulk_insert(
            db,
            task.id,
            [_youtube_candidate_row(task)],
            run_at=datetime.now(UTC),
            product_id=1,
        )
        await db.flush()
        candidate = (
            await db.execute(
                select(CollectionTaskCandidate).where(CollectionTaskCandidate.task_id == task.id)
            )
        ).scalar_one()
        content, _ = build_collection_task_candidates_excel(
            [(candidate, None)],
            task_id=task.id,
            task_name=task.name,
        )
        ws = load_workbook(BytesIO(content)).active
        headers = [cell.value for cell in ws[1]]
        assert ws.cell(row=2, column=headers.index("来源输入链接") + 1).value == AMAZON_NORMALIZED
        assert ws.cell(row=2, column=headers.index("来源作品链接") + 1).value == YOUTUBE_POST
        await db.rollback()


@pytest.mark.anyio
async def test_influencer_library_excel_from_persisted_product_influencer_source():
    from app.models.product_influencer_source import ProductInfluencerSource

    suffix = uuid.uuid4().hex[:8]
    async with async_session_factory() as db:
        task = _competitor_task()
        db.add(task)
        await db.flush()
        global_row = GlobalInfluencerProfile(
            platform="youtube",
            username=f"lib_{suffix}",
            normalized_username=f"lib_{suffix}",
            profile_url=f"https://www.youtube.com/channel/LIB{suffix}",
            normalized_profile_url=f"https://www.youtube.com/channel/LIB{suffix}",
            display_name="Library Creator",
            bio="bio",
            followers_count=1000,
            engagement_rate=2.0,
            email="a@example.com",
            final_email="a@example.com",
        )
        db.add(global_row)
        await db.flush()
        product_row = ProductInfluencer(
            product_id=1,
            global_influencer_id=global_row.id,
        )
        db.add(product_row)
        await db.flush()
        source_row = ProductInfluencerSource(
            product_influencer_id=product_row.id,
            task_id=task.id,
            source_post_url=YOUTUBE_POST,
            source_input_url=AMAZON_NORMALIZED,
            source_platform="youtube",
            task_name=task.name,
            source_key=YOUTUBE_POST.lower(),
            collected_at=datetime.now(UTC),
        )
        db.add(source_row)
        await db.flush()

        from app.services.influencer_projection import merged_influencer_for_ai

        influencer = merged_influencer_for_ai(product_row, global_row)
        sources = await InfluencerSourceService.list_for_product_influencers(db, [product_row.id])
        content, _ = build_influencer_library_excel([influencer], sources_by_influencer_id=sources)
        ws = load_workbook(BytesIO(content)).active
        headers = [cell.value for cell in ws[1]]
        assert ws.cell(row=2, column=headers.index("来源输入链接") + 1).value == AMAZON_NORMALIZED
        assert ws.cell(row=2, column=headers.index("来源作品链接") + 1).value == YOUTUBE_POST
        await db.rollback()
