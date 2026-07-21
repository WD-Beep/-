# 文件说明：后端业务服务，负责采集、筛选、AI、邮件和任务流程；当前文件：export
from datetime import UTC, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models.collection_task_candidate import CollectionTaskCandidate
from app.models.influencer import Influencer
from app.models.product_influencer_source import ProductInfluencerSource
from app.services.contact_signals import build_contact_summary
from app.services.contact_discovery import credibility_level_label
from app.services.value_tier import classify_value_tier, value_tier_label
from app.services.link_seed_enrichment import (
    build_seed_enrichment_status,
    compute_seed_source_export_fields,
    final_platform_display,
    link_seed_platform_display,
    resolve_seed_platform_from_input_url,
)

CANDIDATE_BUSINESS_EXPORT_COLUMNS: list[tuple[str, str, str]] = [
    ("username", "账号", "text"),
    ("display_name", "昵称", "text"),
    ("platform", "平台", "text"),
    ("profile_url", "主页链接", "url"),
    ("source_platform", "来源平台", "text"),
    ("source_input_url", "来源输入链接", "url"),
    ("product_match_reason", "商品匹配说明", "wrap"),
    ("seed_platform", "Seed 平台", "text"),
    ("seed_enrichment_status", "Seed 补全状态", "text"),
    ("followers_count", "粉丝数", "int"),
    ("engagement_rate", "互动率", "float"),
    ("avg_views", "平均观看", "int"),
    ("avg_likes", "平均点赞", "int"),
    ("avg_comments", "平均评论", "int"),
    ("country", "国家", "text"),
    ("language", "语言", "text"),
    ("category", "类目", "text"),
    ("niche", "领域", "text"),
    ("bio", "简介", "wrap"),
    ("email", "邮箱", "text"),
    ("business_email", "商务邮箱", "text"),
    ("public_email", "公开邮箱", "text"),
    ("website", "官网", "text"),
    ("linktree_url", "Linktree/链接页", "text"),
    ("whatsapp", "WhatsApp", "text"),
    ("telegram", "Telegram", "text"),
    ("other_social_links", "其他外链", "wrap"),
    ("contact_summary", "联系方式摘要", "text"),
    ("value_tier", "价值分层", "text"),
    ("value_tier_reason", "分层原因", "wrap"),
    ("recommended_action", "推荐动作", "text"),
    ("score", "综合评分", "float"),
    ("final_priority", "优先级", "text"),
    ("product_fit", "Product Fit", "float"),
    ("engagement_score", "互动分", "float"),
    ("content_match_score", "内容匹配", "float"),
    ("commercial_signal_score", "商业信号", "float"),
    ("contactability_score", "联系方式评分", "float"),
    ("risk_score", "风险分", "float"),
    ("risk_level", "风险等级", "text"),
    ("estimated_collab_price", "预估合作价格", "text"),
    ("roi_forecast", "ROI 预估", "float"),
    ("collaboration_formats", "合作形式", "text"),
    ("content_topics", "内容标签", "text"),
    ("ai_summary", "AI 推荐理由", "wrap"),
    ("ai_collaboration_suggestion", "适合怎么合作", "wrap"),
    ("ai_outreach_message", "开场话术", "wrap"),
    ("source_post_url", "来源作品链接", "url"),
    ("source_task_name", "来源任务", "text"),
    ("collected_at", "采集时间", "text"),
]

INFLUENCER_LIBRARY_EXPORT_COLUMNS: list[tuple[str, str, str]] = [
    ("username", "账号", "text"),
    ("display_name", "昵称", "text"),
    ("platform", "平台", "text"),
    ("profile_url", "主页链接", "url"),
    ("source_platform", "来源平台", "text"),
    ("source_input_url", "来源输入链接", "url"),
    ("seed_platform", "Seed 平台", "text"),
    ("seed_enrichment_status", "Seed 补全状态", "text"),
    ("followers_count", "粉丝数", "int"),
    ("engagement_rate", "互动率", "float"),
    ("email", "邮箱", "text"),
    ("bio", "简介", "wrap"),
    ("source_post_url", "来源作品链接", "url"),
    ("source_task_name", "来源任务", "text"),
    ("collected_at", "采集时间", "text"),
]

EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("platform", "平台"),
    ("display_name", "昵称"),
    ("profile_url", "主页链接"),
    ("source_platform", "来源平台"),
    ("source_input_url", "来源输入链接"),
    ("seed_platform", "Seed 平台"),
    ("seed_enrichment_status", "Seed 补全状态"),
    ("source_discovery_type", "发现来源"),
    ("source_post_url", "来源作品链接"),
    ("source_task_name", "来源任务"),
    ("collected_at", "采集时间"),
    ("source_comment_text", "来源评论"),
    ("public_email", "公开邮箱"),
    ("business_email", "商业邮箱"),
    ("contact_page", "联系页"),
    ("linktree_url", "Linktree/链接页"),
    ("other_social_links", "其他外链"),
    ("contact_score", "联系方式评分"),
    ("contact_credibility_level", "联系可信度"),
    ("value_tier", "价值分层"),
    ("value_tier_reason", "分层原因"),
    ("country", "国家"),
    ("language", "语言"),
    ("category", "类目"),
    ("followers_count", "粉丝数"),
    ("avg_views", "平均观看"),
    ("avg_likes", "平均点赞"),
    ("avg_comments", "平均评论"),
    ("engagement_rate", "互动率"),
    ("final_priority", "优先级"),
    ("engagement_score", "互动分"),
    ("content_match_score", "内容匹配分"),
    ("contactability_score", "可联系分"),
    ("commercial_signal_score", "商业信号分"),
    ("activity_score", "活跃度分"),
    ("risk_score", "风险分"),
    ("score", "综合评分"),
    ("product_fit", "产品匹配度"),
    ("travel_fit_score", "Travel Fit Score"),
    ("purchasing_power_score", "购买力评分"),
    ("sales_potential_score", "带货能力评分"),
    ("audience_match_score", "受众匹配度"),
    ("roi_forecast", "ROI 预估"),
    ("email", "邮箱"),
    ("email_source", "邮箱来源"),
    ("website", "网站"),
    ("ai_summary", "AI 画像"),
    ("ai_collaboration_suggestion", "合作建议"),
    ("ai_outreach_message", "触达话术"),
    ("score_reason", "评分理由"),
]


def _cell_value(
    influencer: Influencer,
    field: str,
    *,
    source_fields: dict[str, str] | None = None,
):
    if source_fields and field in source_fields:
        return source_fields[field]
    if field == "display_name":
        return influencer.display_name or influencer.username or ""
    if field == "platform":
        value = getattr(influencer, "platform", None)
        return final_platform_display(value) or value or ""
    if field == "value_tier":
        tier, _, _ = classify_value_tier(influencer)
        return value_tier_label(tier)
    if field == "value_tier_reason":
        _, _, reason = classify_value_tier(influencer)
        return reason
    value = getattr(influencer, field, None)
    if value is None:
        return ""
    if field == "contact_credibility_level":
        return credibility_level_label(str(value))
    if field == "last_collected_at" and isinstance(value, datetime):
        return value.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S")
    if field == "engagement_rate" and isinstance(value, float):
        return f"{value:.2f}%"
    if field == "roi_forecast" and isinstance(value, float):
        return f"{value:.1f}x"
    if field == "other_social_links":
        return _format_social_links(value if isinstance(value, list) else None)
    return value


def _resolve_export_email(influencer: Influencer) -> str:
    for field in ("final_email", "email", "public_email", "business_email"):
        value = getattr(influencer, field, None)
        if value:
            return str(value).strip()
    return ""


def _library_export_value(
    influencer: Influencer,
    field: str,
    *,
    source_fields: dict[str, str] | None = None,
) -> str:
    if field in {
        "source_post_url",
        "source_input_url",
        "source_task_name",
        "collected_at",
        "source_platform",
        "seed_platform",
        "seed_enrichment_status",
    }:
        if source_fields:
            return source_fields.get(field, "")
        if field == "source_post_url":
            return str(getattr(influencer, "source_post_url", None) or "").strip()
        return ""
    if field == "profile_url":
        return str(getattr(influencer, "profile_url", None) or "").strip()
    if field == "email":
        return _resolve_export_email(influencer)
    if field == "display_name":
        return str(getattr(influencer, "display_name", None) or getattr(influencer, "username", None) or "").strip()
    if field == "username":
        return str(getattr(influencer, "username", None) or "").strip()
    if field == "platform":
        return final_platform_display(getattr(influencer, "platform", None)) or str(
            getattr(influencer, "platform", None) or ""
        ).strip()
    if field == "bio":
        return str(getattr(influencer, "bio", None) or "").strip()
    if field == "followers_count":
        value = getattr(influencer, "followers_count", None)
        return "" if value is None else str(value)
    if field == "engagement_rate":
        value = getattr(influencer, "engagement_rate", None)
        if value is None:
            return ""
        return f"{float(value):.2f}%"
    return ""


def _write_library_export_cell(ws, row_idx: int, col_idx: int, kind: str, raw_value: str) -> None:
    cell = ws.cell(row=row_idx, column=col_idx)
    link_font = Font(color="0563C1", underline="single")

    if not raw_value:
        cell.value = None
        return

    if kind == "url":
        cell.value = raw_value
        cell.hyperlink = raw_value
        cell.font = link_font
        return

    cell.value = raw_value


def _auto_fit_library_columns(ws, row_count: int) -> None:
    widths = {
        "主页链接": 42,
        "邮箱": 28,
        "来源作品链接": 42,
        "来源输入链接": 42,
        "Seed 补全状态": 36,
        "简介": 36,
    }
    for col_idx in range(1, len(INFLUENCER_LIBRARY_EXPORT_COLUMNS) + 1):
        header = str(ws.cell(row=1, column=col_idx).value or "")
        column_letter = get_column_letter(col_idx)
        if header in widths:
            ws.column_dimensions[column_letter].width = widths[header]
            continue
        max_length = len(header)
        for row_idx in range(2, row_count + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def build_influencer_library_excel(
    influencers: list[Influencer] | list,
    filename: str | None = None,
    *,
    sources_by_influencer_id: dict[int, list[ProductInfluencerSource]] | None = None,
) -> tuple[bytes, str]:
    from app.services.influencer_source import InfluencerSourceService

    wb = Workbook()
    ws = wb.active
    ws.title = "红人数据"

    headers = [label for _, label, _ in INFLUENCER_LIBRARY_EXPORT_COLUMNS]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for row_offset, influencer in enumerate(influencers, start=2):
        source_fields = None
        source_rows: list | None = None
        influencer_id = getattr(influencer, "id", None)
        if sources_by_influencer_id and influencer_id is not None:
            source_rows = sources_by_influencer_id.get(influencer_id, [])
            source_fields = InfluencerSourceService.aggregate_for_export(source_rows)
            source_fields = _merge_source_export_fields(
                source_fields,
                sources=source_rows,
                final_platform=getattr(influencer, "platform", None),
            )
        for col_idx, (field, _, kind) in enumerate(INFLUENCER_LIBRARY_EXPORT_COLUMNS, start=1):
            raw_value = _library_export_value(influencer, field, source_fields=source_fields)
            _write_library_export_cell(ws, row_offset, col_idx, kind, raw_value)

    ws.freeze_panes = "A2"
    _auto_fit_library_columns(ws, len(influencers))

    filename = filename or f"influencer_export_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename


def _join_list_field(value: list | None) -> str:
    if not value:
        return ""
    return ", ".join(str(item) for item in value if item)


def _format_social_links(value: list | None) -> str:
    if not value:
        return ""
    parts: list[str] = []
    for link in value:
        if not isinstance(link, dict):
            continue
        url = str(link.get("url") or "").strip()
        if not url:
            continue
        label = str(link.get("label") or link.get("type") or "外链").strip()
        parts.append(f"{label}: {url}")
    return "\n".join(parts)


def _candidate_source_input_url(candidate: CollectionTaskCandidate) -> str:
    column = getattr(candidate, "source_input_url", None)
    if column and str(column).strip():
        return str(column).strip()
    meta = getattr(candidate, "source_meta", None) or {}
    if isinstance(meta, dict):
        value = meta.get("source_input_url") or meta.get("input_url")
        if value:
            return str(value).strip()
    return ""


def _candidate_link_seed_meta(candidate: CollectionTaskCandidate) -> dict:
    meta = getattr(candidate, "source_meta", None) or {}
    if not isinstance(meta, dict):
        return {}
    enrichment = meta.get("link_seed_enrichment")
    return enrichment if isinstance(enrichment, dict) else {}


def _candidate_final_platform(
    candidate: CollectionTaskCandidate,
    influencer: Influencer | None,
) -> str:
    if influencer and getattr(influencer, "platform", None):
        return str(influencer.platform)
    return str(getattr(candidate, "platform", None) or "")


def _merge_source_export_fields(
    source_fields: dict[str, str] | None,
    *,
    sources: list | None,
    final_platform: str | None,
) -> dict[str, str]:
    merged = dict(source_fields or {})
    seed_fields = compute_seed_source_export_fields(sources, final_platform=final_platform)
    if seed_fields.get("source_platform"):
        merged["source_platform"] = seed_fields["source_platform"]
    elif merged.get("source_platform"):
        merged["source_platform"] = "\n".join(
            final_platform_display(part.strip()) or part.strip()
            for part in merged["source_platform"].split("\n")
            if part.strip()
        )
    merged["seed_platform"] = seed_fields.get("seed_platform", "")
    merged["seed_enrichment_status"] = seed_fields.get("seed_enrichment_status", "")
    return merged


def _candidate_export_value(
    candidate: CollectionTaskCandidate,
    influencer: Influencer | None,
    field: str,
    *,
    task_name: str | None = None,
    task_id: int | None = None,
):
    if field == "username":
        return candidate.username or (influencer.username if influencer else "") or ""

    if field == "platform":
        final = _candidate_final_platform(candidate, influencer)
        display = final_platform_display(final)
        return display or final

    if field == "profile_url":
        if influencer and influencer.profile_url:
            return influencer.profile_url
        return candidate.profile_url or ""

    if field == "source_post_url":
        return candidate.source_post_url or (influencer.source_post_url if influencer else None) or ""

    if field == "source_input_url":
        return _candidate_source_input_url(candidate)

    if field == "product_match_reason":
        meta = getattr(candidate, "source_meta", None) or {}
        if not isinstance(meta, dict):
            return ""
        selected = str(meta.get("selected_reason") or "").strip()
        if selected:
            return selected
        reasons = meta.get("match_reasons")
        if isinstance(reasons, list):
            return "；".join(str(item) for item in reasons if item)
        return ""

    if field == "source_task_name":
        return task_name or ""

    if field == "source_platform":
        enrichment = _candidate_link_seed_meta(candidate)
        seed_key = str(enrichment.get("link_seed_platform") or "").strip().lower()
        if seed_key:
            return link_seed_platform_display(seed_key)
        inp = _candidate_source_input_url(candidate)
        detected = resolve_seed_platform_from_input_url(inp)
        if detected:
            return link_seed_platform_display(detected)
        cand_plat = str(getattr(candidate, "platform", None) or "").strip().lower()
        if cand_plat in {"ltk", "shopmy", "pinterest"}:
            return link_seed_platform_display(cand_plat)
        return ""

    if field == "seed_platform":
        enrichment = _candidate_link_seed_meta(candidate)
        seed_key = str(enrichment.get("link_seed_platform") or "").strip().lower()
        if not seed_key:
            seed_key = resolve_seed_platform_from_input_url(_candidate_source_input_url(candidate)) or ""
        return link_seed_platform_display(seed_key) if seed_key else ""

    if field == "seed_enrichment_status":
        enrichment = _candidate_link_seed_meta(candidate)
        return build_seed_enrichment_status(
            enrichment_meta=enrichment or None,
            source_input_url=_candidate_source_input_url(candidate),
            final_platform=_candidate_final_platform(candidate, influencer),
            failure_reason=getattr(candidate, "failure_reason", None),
        )

    if field == "collected_at":
        run_at = getattr(candidate, "run_at", None) or getattr(candidate, "profile_fetched_at", None)
        if run_at and isinstance(run_at, datetime):
            return run_at.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S")
        created = getattr(candidate, "created_at", None)
        if created and isinstance(created, datetime):
            return created.astimezone(UTC).strftime("%Y-%m-%d %H:%M:%S")
        return ""

    if field == "display_name":
        if influencer:
            return influencer.display_name or influencer.username or ""
        return ""

    if field in {"followers_count", "engagement_rate"}:
        if influencer is not None:
            value = getattr(influencer, field, None)
            if value is not None:
                return value
        return getattr(candidate, field, None)

    if field == "collaboration_formats":
        return _join_list_field(influencer.collaboration_formats if influencer else None)

    if field == "content_topics":
        if influencer:
            return _join_list_field(influencer.content_topics or influencer.tags)
        return ""

    if field == "other_social_links":
        return _format_social_links(influencer.other_social_links if influencer else None)

    if field == "contact_summary":
        if influencer:
            return build_contact_summary(influencer)
        return ""

    if field == "value_tier":
        if influencer:
            tier, label, _ = classify_value_tier(influencer)
            return label or value_tier_label(tier)
        return ""

    if field == "value_tier_reason":
        if influencer:
            _, _, reason = classify_value_tier(influencer)
            return reason
        return ""

    if field == "recommended_action":
        if not influencer:
            if candidate.status == "filtered_out":
                return "跳过"
            return "待入库后评估"
        tier, _, _ = classify_value_tier(influencer)
        if tier == "direct_contact":
            return "直接联系"
        if tier == "manual_research":
            return "人工补充联系方式"
        return "跳过"

    if field == "ai_summary":
        if influencer:
            return influencer.ai_summary or influencer.score_reason or ""
        return ""

    if influencer is not None:
        return getattr(influencer, field, None)
    return None


def _write_candidate_export_cell(ws, row_idx: int, col_idx: int, field: str, kind: str, raw_value) -> None:
    cell = ws.cell(row=row_idx, column=col_idx)
    link_font = Font(color="0563C1", underline="single")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    if raw_value is None or raw_value == "":
        cell.value = None
        return

    if kind == "url":
        url = str(raw_value).strip()
        cell.value = url
        cell.hyperlink = url
        cell.font = link_font
        return

    if kind == "int":
        cell.value = int(raw_value)
        cell.number_format = "0"
        return

    if kind == "float":
        cell.value = float(raw_value)
        cell.number_format = "0.00"
        return

    if kind == "wrap":
        cell.value = str(raw_value)
        cell.alignment = wrap_align
        return

    cell.value = str(raw_value)


def _auto_fit_candidate_columns(ws, row_count: int, col_count: int) -> None:
    wrap_labels = {label for _, label, kind in CANDIDATE_BUSINESS_EXPORT_COLUMNS if kind == "wrap"}
    url_labels = {label for _, label, kind in CANDIDATE_BUSINESS_EXPORT_COLUMNS if kind == "url"}

    for col_idx in range(1, col_count + 1):
        column_letter = get_column_letter(col_idx)
        header = str(ws.cell(row=1, column=col_idx).value or "")
        if header in wrap_labels:
            ws.column_dimensions[column_letter].width = 36
            continue
        if header in url_labels:
            ws.column_dimensions[column_letter].width = 42
            continue

        max_length = len(header)
        for row_idx in range(2, row_count + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 28)


def build_collection_task_candidates_excel(
    rows: list[tuple[CollectionTaskCandidate, Influencer | None]],
    *,
    task_id: int,
    task_name: str | None = None,
) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "候选池"

    headers = [label for _, label, _ in CANDIDATE_BUSINESS_EXPORT_COLUMNS]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for row_offset, (candidate, influencer) in enumerate(rows, start=2):
        for col_idx, (field, _, kind) in enumerate(CANDIDATE_BUSINESS_EXPORT_COLUMNS, start=1):
            raw_value = _candidate_export_value(
                candidate,
                influencer,
                field,
                task_name=task_name,
                task_id=task_id,
            )
            _write_candidate_export_cell(ws, row_offset, col_idx, field, kind, raw_value)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    last_row = max(len(rows) + 1, 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    _auto_fit_candidate_columns(ws, len(rows), len(CANDIDATE_BUSINESS_EXPORT_COLUMNS))

    filename = f"collection-task-{task_id}-candidates.xlsx"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename


def _auto_fit_columns(ws, row_count: int, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        column_letter = get_column_letter(col_idx)
        header = str(ws.cell(row=1, column=col_idx).value or "")
        max_length = len(header)

        for row_idx in range(2, row_count + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def build_influencer_excel(
    influencers: list[Influencer] | list,
    filename: str | None = None,
    *,
    sources_by_influencer_id: dict[int, list[ProductInfluencerSource]] | None = None,
) -> tuple[bytes, str]:
    from app.services.influencer_source import InfluencerSourceService

    wb = Workbook()
    ws = wb.active
    ws.title = "红人数据"

    headers = [label for _, label in EXPORT_COLUMNS]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    for influencer in influencers:
        source_fields = None
        source_rows: list | None = None
        influencer_id = getattr(influencer, "id", None)
        if sources_by_influencer_id and influencer_id is not None:
            source_rows = sources_by_influencer_id.get(influencer_id, [])
            source_fields = InfluencerSourceService.aggregate_for_export(source_rows)
            source_fields = _merge_source_export_fields(
                source_fields,
                sources=source_rows,
                final_platform=getattr(influencer, "platform", None),
            )
        row = [
            _cell_value(influencer, field, source_fields=source_fields)
            for field, _ in EXPORT_COLUMNS
        ]
        ws.append(row)

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws, len(influencers), len(EXPORT_COLUMNS))

    filename = filename or f"influencer_export_{datetime.now(UTC).strftime('%Y%m%d_%H%M')}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), filename
