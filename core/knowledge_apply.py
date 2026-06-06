"""将以 LLM 裁决通过的新物料行写入 price_kb.xlsx。

写入判重（仅此文件，不动闭环其它环节）：仅当「名称+规格+单价」经规范化后与表中某行**完全一致**时跳过；
禁止模糊/向量/语义去重。前三列为业务列，第四列用文字标记自动补录进知识库（与拆解里验算的 ✓ 不同）。
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from embedding.embedding_index import get_embedding_index
from price_kb import note_kb_disk_write_success

# 自动学习写入行的第四列标记（不影响 price_kb 三列解析；勿用单独的 √，易与拆解验算勾选混淆）
AUTO_LEARN_ROW_MARKER = "KB自增"


def _pick_material_sheet(open_workbook: Any):
    wb = open_workbook
    for sn in wb.sheetnames:
        if "询价" in sn or "材料" in sn:
            return wb[sn]
    return wb[wb.sheetnames[0]]


def _ensure_auto_mark_column_header(ws: Any) -> None:
    """若首行像表头且 D 列尚无标题，补上「知识库自增」列名，便于识别补录标记列。"""
    try:
        v1 = ws.cell(row=1, column=1).value
        v4 = ws.cell(row=1, column=4).value
        h = str(v1 or "")
        if ("材料" not in h and "名称" not in h) or str(v4 or "").strip():
            return
        ws.cell(row=1, column=4, value="知识库自增")
    except Exception:
        return


# 严格增量写入：仅「规范化后名称+规格+单价」完全一致才视为重复；禁止模糊/语义/向量去重。
_WS_COLLAPSE = re.compile(r"\s+")


def _normalize_kb_cell(value: object) -> str:
    """规范化单格：NFC、trim、小写、空白压成单空格（用于 canonical exact 比较）。"""
    t = unicodedata.normalize("NFC", str(value or ""))
    t = t.strip().lower()
    t = _WS_COLLAPSE.sub(" ", t)
    return t.strip()


def _canonical_material_triple(name: str, spec: str, price: str) -> tuple[str, str, str]:
    return (
        _normalize_kb_cell(name),
        _normalize_kb_cell(spec if str(spec).strip() else "-"),
        _normalize_kb_cell(price),
    )


def _has_kb_garbage_symbol(*values: object) -> bool:
    return any("?" in str(value or "") or "？" in str(value or "") for value in values)


def _material_row_is_safe_to_write(name: str, spec: str, price: str) -> bool:
    if not str(name or "").strip() or not str(price or "").strip():
        return False
    if _has_kb_garbage_symbol(name, spec, price):
        return False
    from kb_data_quality import KB_ACTION_AUTO, judge_kb_insert_candidate

    verdict = judge_kb_insert_candidate(name, spec, price)
    return verdict.action == KB_ACTION_AUTO


def _canonical_row_already_exists(ws: Any, wanted: tuple[str, str, str]) -> bool:
    """表中任一行规范化三元组与 wanted 完全一致则视为已有，不追加。"""
    if not wanted[0]:
        return False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        cells = [str(c or "") for c in row[:3]]
        if not cells or not str(cells[0]).strip():
            continue
        rn = cells[0]
        rs = cells[1] if len(cells) > 1 else ""
        rp = cells[2] if len(cells) > 2 else ""
        ext = _canonical_material_triple(rn, rs, rp)
        if ext == wanted:
            return True
    return False


def kb_material_row_exists(material: dict[str, Any], kb_path: Path) -> bool:
    """Return True when the canonical [name, spec, price] triple already exists."""
    name = str(material.get("name") or "").strip()
    spec = str(material.get("spec") or "").strip() or "-"
    price = str(material.get("price") or material.get("unit_price") or "").strip()
    if not _material_row_is_safe_to_write(name, spec, price):
        return False

    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    path = kb_path.resolve()
    if not path.is_file():
        return False

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _pick_material_sheet(wb)
        return _canonical_row_already_exists(ws, _canonical_material_triple(name, spec, price))
    finally:
        try:
            wb.close()
        except Exception:
            pass


def kb_material_name_spec_exists(material: dict[str, Any], kb_path: Path) -> bool:
    """Return True when canonical [name, spec] already exists, regardless of price."""
    name = str(material.get("name") or "").strip()
    spec = str(material.get("spec") or "").strip() or "-"
    if not str(name or "").strip():
        return False
    if _has_kb_garbage_symbol(name, spec):
        return False

    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    path = kb_path.resolve()
    if not path.is_file():
        return False

    wanted_name = _normalize_kb_cell(name)
    wanted_spec = _normalize_kb_cell(spec)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _pick_material_sheet(wb)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            row_name = str((row[0] if len(row) > 0 else "") or "").strip()
            row_spec = str((row[1] if len(row) > 1 else "") or "").strip() or "-"
            if not row_name:
                continue
            if _normalize_kb_cell(row_name) == wanted_name and _normalize_kb_cell(row_spec) == wanted_spec:
                return True
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass


def apply_kb_write(material: dict[str, Any], kb_path: Path) -> bool:
    """向工作表追加一行 [材料名称, 规格, 单价, KB自增]。调用方已持 KNOWLEDGE_MUTATION_LOCK。"""
    from price_kb_paths import assert_official_kb_write_allowed, is_official_kb_path

    path = Path(kb_path).expanduser().resolve()
    if is_official_kb_path(path):
        try:
            assert_official_kb_write_allowed(path, updated_by="agent_auto", source="knowledge_auto")
        except PermissionError as exc:
            print(f"[knowledge-apply] blocked official write: {exc}", flush=True)
            return False

    name = str(material.get("name") or "").strip()
    spec = str(material.get("spec") or "").strip() or "-"
    price = str(material.get("price") or material.get("unit_price") or "").strip()
    if not _material_row_is_safe_to_write(name, spec, price):
        print("[knowledge-apply] skip: unsafe or incomplete material row", flush=True)
        return False

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[knowledge-apply] 需要安装 openpyxl：pip install openpyxl", flush=True)
        return False

    if not path.is_file():
        print(f"[knowledge-apply] missing file {path}", flush=True)
        return False

    try:
        wb = load_workbook(path)
        ws = _pick_material_sheet(wb)
        want_key = _canonical_material_triple(name, spec, price)
        if _canonical_row_already_exists(ws, want_key):
            print(
                f"[knowledge-apply] strict-dedupe skip (canonical triple match): "
                f"name={name!r} spec={spec!r}",
                flush=True,
            )
            return False
        _ensure_auto_mark_column_header(ws)
        ws.append([name, spec, price, AUTO_LEARN_ROW_MARKER])
        wb.save(path)
        # 磁盘已为新版本：立即禁用旧索引，直至 knowledge_reload_hook→prepare 完成
        get_embedding_index().mark_unready()
        note_kb_disk_write_success(path)
    except OSError as exc:
        print(f"[knowledge-apply] save failed: {exc}", flush=True)
        return False

    print(f"[knowledge-apply] appended row name={name!r} spec={spec!r}", flush=True)
    return True
