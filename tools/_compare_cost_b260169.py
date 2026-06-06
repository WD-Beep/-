# -*- coding: utf-8
import sys, base64, json
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from demand_parser import parse_demand_from_payload, compute_mold_fee_from_sections
from quote_engine import calculate_quote, build_settings
from structure_usage import apply_structure_usage_hints, tighten_small_bag_usage_amounts
from price_kb import get_price_kb

XLSX = Path(r"d:\测试数据\B260169--报价资料.xlsx")
b64 = base64.b64encode(XLSX.read_bytes()).decode()
parsed = parse_demand_from_payload({"name": XLSX.name, "content_base64": b64})
kb = get_price_kb()

items = []
for m in parsed.materials:
    row = {"name": m.name, "spec": m.spec or "-", "usage": "-", "unit_price": m.inline_price or "-", "amount": 0.0}
    if m.inline_price:
        row["unit_price_ai"] = True
    hit = kb.lookup(m.name, m.spec or "")
    if hit and not m.inline_price:
        row["unit_price"] = hit.entry.raw_price
    items.append(row)

st = parsed.structure_text or ""
apply_structure_usage_hints(items, st, product_size=parsed.product_size)
tighten_small_bag_usage_amounts(items, product_size=parsed.product_size, structure_text=st)

# 用 build_settings 计算 amount
settings = build_settings({"items": items})
items_with_amt = [it.to_dict() for it in settings.items]

payload = {
    "items": items_with_amt,
    "quantities": list(parsed.quantities) if parsed.quantities else [300, 500, 1000],
    "product_name": parsed.product_name,
    "mold_fee": float(compute_mold_fee_from_sections(parsed.sections)),
    "include_fob": True,
}
qs = parsed.quote_settings
if qs.get("processing_fee") is not None:
    payload["processing_fee"] = float(qs["processing_fee"])
if qs.get("management_loss_rate") is not None:
    payload["management_loss_rate"] = float(qs["management_loss_rate"])
if qs.get("system_overhead_fixed") is not None:
    payload["system_overhead_fixed"] = float(qs["system_overhead_fixed"])
if qs.get("gross_margin_rate") is not None:
    payload["gross_margin_rate"] = float(qs["gross_margin_rate"])

result = calculate_quote(payload)

print("=== 需求表 quote_settings ===")
for k in ("processing_fee", "processing_fee_rule", "processing_fee_locked", "processing_fee_cap", "management_loss_rate", "system_overhead_fixed"):
    print(f"  {k}: {qs.get(k)}")

print("\n=== 物料明细 (Agent 路径) ===")
mat = 0.0
for it in items_with_amt:
    a = float(it.get("amount") or 0)
    mat += a
    print(f"  {it['name']:14} usage={str(it.get('usage','')):10} price={str(it.get('unit_price','')):12} sub={a:.2f}")
print(f"  物料合计: {result.get('material_total')} (手加 {mat:.2f})")

print("\n=== 成本公式组成 ===")
print(f"  mold_fee总额: {result.get('mold_fee')}")
print(f"  processing_fee: {result.get('processing_fee')}")
print(f"  system_overhead: {result.get('system_overhead')} ({result.get('system_overhead_rule')})")

print("\n=== 各档 毛利前成本 = 物料 + 管理费 + 加工费 + 刀模摊 ===")
for t in result.get("tiers") or []:
    q = t.get("quantity")
    cbm = t.get("cost_before_margin")
    ms = t.get("mold_share")
    pf = t.get("processing_fee")
    print(f"  {q}件: {cbm} = 物料{result.get('material_total')} + 管理费{result.get('system_overhead')} + 加工{pf} + 刀模摊{ms}")

print("\n=== 对照业务员 (500件总成本 47.11) ===")
print("  业务员: 小计不含刀模 45.11 + 刀模摊 2.00 = 47.11")
print("  业务员 BOM 物料小计约 25.82 (图二/三)")
t500 = next((t for t in result.get("tiers") or [] if int(t.get("quantity") or 0) == 500), None)
if t500:
    agent = float(t500.get("cost_before_margin") or 0)
    print(f"  Agent 500件: {agent}  差额 +{round(agent - 47.11, 2)}")
    print(f"  拆解差额约: 物料多 +{round(float(result.get('material_total') or 0) - 25.82, 2)}")
    pf = float(result.get("processing_fee") or 0)
    oh = float(result.get("system_overhead") or 0)
    print(f"  Agent 加工费={pf} 管理费={oh}  (45.11-25.82=19.29 为业务员非物料+刀模前小计)")

# R33 成本字段
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb[wb.sheetnames[0]]
for r in range(30, 36):
    vals = [str(ws.cell(r,c).value) for c in range(1,8) if ws.cell(r,c).value]
    if vals: print(f"  需求表R{r}: {' | '.join(vals)}")
