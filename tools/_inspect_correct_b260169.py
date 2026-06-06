# -*- coding: utf-8
import sys
import base64
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from demand_parser import (
    parse_demand_from_payload,
    quotation_detail_materials_bundle_from_entire_xlsx,
    collect_quotation_detail_materials_from_xlsx,
    collect_auxiliary_bom_materials_from_xlsx,
)
from quotation_detail_table import find_quotation_detail_header_row
from simple_bom_parser import is_simple_bom_template

XLSX = Path(r"d:\正确数\B260169.xlsx")

wb = openpyxl.load_workbook(XLSX, data_only=True)
print("=" * 70)
print("1. 工作表列表")
print("=" * 70)
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"  [{i}] {name!r}  rows={ws.max_row} cols={ws.max_column}")

print("\n" + "=" * 70)
print("2. 各表前8行 + 关键行扫描")
print("=" * 70)
keywords = ("物料", "BOM", "清单", "辅料", "面料", "小计", "总成本", "加工", "刀模", "45.", "47.", "25.", "需求表", "填写")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n--- {name} ---")
    for r in range(1, min(9, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value).strip()[:28] for c in range(1, min(9, ws.max_column + 1)) if ws.cell(r, c).value is not None]
        if vals:
            print(f"  R{r}: {' | '.join(vals)}")
    for r in range(1, ws.max_row + 1):
        joined = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(10, ws.max_column + 1)))
        if any(k in joined for k in keywords):
            if len(joined) > 120:
                joined = joined[:120] + "..."
            print(f"  *R{r}: {joined}")

print("\n" + "=" * 70)
print("3. 系统识别能力")
print("=" * 70)
b64 = base64.b64encode(XLSX.read_bytes()).decode()

# 需求表
try:
    parsed = parse_demand_from_payload({"name": XLSX.name, "content_base64": b64})
    print(f"需求表解析: OK  product={parsed.product_name!r} materials={len(parsed.materials)}")
    print(f"  auxiliary_bom_sheets: {parsed.auxiliary_bom_sheet_names}")
    print(f"  processing_fee={parsed.quote_settings.get('processing_fee')} rule={parsed.quote_settings.get('processing_fee_rule')}")
except Exception as e:
    print(f"需求表解析: FAIL {e}")

# 报价明细 bundle
bundle = quotation_detail_materials_bundle_from_entire_xlsx(XLSX.read_bytes())
print(f"全簿报价明细 bundle: {len(bundle)} 条")
for m in bundle[:12]:
    print(f"  {m.name!r} usage={m.quoted_usage!r} calc={str(m.calc_method or '')[:50]}")

# 辅 sheet BOM
detail_ex, detail_names = collect_quotation_detail_materials_from_xlsx(XLSX.read_bytes(), "", file_name=XLSX.name)
simple_ex, simple_names = collect_auxiliary_bom_materials_from_xlsx(XLSX.read_bytes(), "", file_name=XLSX.name)
print(f"collect_quotation_detail: {len(detail_ex)}  sheets={detail_names}")
print(f"collect_auxiliary_bom: {len(simple_ex)}  sheets={simple_names}")

# 每张 sheet 检测
for name in wb.sheetnames:
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, min(ws.max_row, 200) + 1)]
    rows = [["" if c is None else c for c in row] for row in rows]
    hi = find_quotation_detail_header_row(rows)
    bom = is_simple_bom_template(rows)
    print(f"  sheet {name!r}: quotation_detail_header={hi} simple_bom={bom}")

print("\n" + "=" * 70)
print("4. 业务员成本关键数（扫描）")
print("=" * 70)
for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 12)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v)
            if any(x in s for x in ("45.11", "47.11", "25.82", "60.47", "60.51", "小计", "总成本")):
                row = [str(ws.cell(r, cc).value)[:20] for cc in range(1, 8) if ws.cell(r, cc).value is not None]
                print(f"  [{name}] R{r}: {' | '.join(row)}")
