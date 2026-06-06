# -*- coding: utf-8 -*-
import sys
import base64
import json
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from demand_parser import parse_demand_from_payload, quotation_detail_materials_bundle_from_entire_xlsx
from quotation_detail_table import find_quotation_detail_header_row

XLSX = Path(r"d:\测试数据\B260169--报价资料.xlsx")

wb = openpyxl.load_workbook(XLSX, data_only=True)
main_name = wb.sheetnames[0]
rows = [["" if c is None else c for c in r] for r in wb[main_name].iter_rows(values_only=True)]

print("=" * 60)
print("1. 工作簿结构")
print("=" * 60)
for n in wb.sheetnames:
    print(" ", n)
print("内嵌报价明细表头行:", find_quotation_detail_header_row(rows))

b64 = base64.b64encode(XLSX.read_bytes()).decode()
parsed = parse_demand_from_payload({"name": XLSX.name, "content_base64": b64})
bundle = quotation_detail_materials_bundle_from_entire_xlsx(XLSX.read_bytes())

print("\n" + "=" * 60)
print("2. 需求表解析出的物料行 (Agent BOM 输入)")
print("=" * 60)
print(f"产品: {parsed.product_name}  尺寸: {parsed.product_size}")
print(f"物料数: {len(parsed.materials)}  全簿报价明细 bundle: {len(bundle)} 条")
for m in parsed.materials:
    print(
        f"  [{m.role:4}] {m.name:16} | spec={str(m.spec or '-'):8} | "
        f"quoted_usage={str(m.quoted_usage or '-'):8} | inline_price={str(m.inline_price or '-'):10} | "
        f"calc={str(m.calc_method or '')[:50]}"
    )

print("\n" + "=" * 60)
print("3. C 区表单字段 (下拉/填写)")
print("=" * 60)
sec_c = parsed.sections.get("C") or {}
for k, v in sorted(sec_c.items()):
    if str(v).strip():
        print(f"  {k}: {v}")

print("\n" + "=" * 60)
print("4. 结构说明中的参考价/用量片段")
print("=" * 60)
st = parsed.structure_text or ""
for chunk in st.replace("\r", "\n").split("\n"):
    t = chunk.strip()
    if t and any(x in t for x in ("元", "码", "米", "X-PAC", "尼龙", "拉链", "织带", "扣")):
        print(" ", t[:100])

print("\n" + "=" * 60)
print("5. 扫描主表是否含「物料清单/BOM/辅料类」区块")
print("=" * 60)
for i, row in enumerate(rows, 1):
    joined = " ".join(str(x) for x in row if x is not None).strip()
    if any(k in joined for k in ("物料清单", "BOM", "辅料类", "面料类", "物料小计", "主体面料", "0.2096")):
        print(f"  R{i}: {joined[:120]}")
