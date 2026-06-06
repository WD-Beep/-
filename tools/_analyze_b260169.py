# -*- coding: utf-8 -*-
import sys
from pathlib import Path
sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
from demand_parser import parse_demand_from_payload, quotation_detail_materials_bundle_from_entire_xlsx
from quotation_detail_table import find_quotation_detail_header_row

p = Path(r"d:\测试数据\B260169--报价资料.xlsx")
wb = openpyxl.load_workbook(p, data_only=True)
print("=== 工作表 ===")
for n in wb.sheetnames:
    print(" ", n)

main = wb.sheetnames[0]
rows = [["" if c is None else c for c in r] for r in wb[main].iter_rows(values_only=True)]
print(f"\n=== {main} 内嵌报价明细表头 ===")
print("header row:", find_quotation_detail_header_row(rows))

print("\n=== C区/结构说明 片段 ===")
for i, row in enumerate(rows, 1):
    line = " | ".join(str(x).strip() for x in row[:6] if str(x).strip())
    if any(k in line for k in ("结构说明", "拉链", "X-PAC", "物料清单", "BOM", "二、")):
        print(f"R{i}: {line[:120]}")

import base64
b64 = base64.b64encode(p.read_bytes()).decode()
parsed = parse_demand_from_payload({"name": p.name, "content_base64": b64})
print(f"\n=== demand 解析: 产品={parsed.product_name} 物料数={len(parsed.materials)} ===")
for m in parsed.materials:
    print(f"  [{m.role}] {m.name} spec={m.spec!r} quoted_usage={m.quoted_usage!r} calc={str(m.calc_method or '')[:60]}")

bundle = quotation_detail_materials_bundle_from_entire_xlsx(p.read_bytes())
print(f"\n=== xlsx 全簿报价明细 bundle: {len(bundle)} 条 ===")
for m in bundle[:15]:
    print(f"  {m.name} usage={m.quoted_usage!r} price={m.inline_price or m.calc_method[:30] if m.calc_method else ''}")

print("\n=== structure_text 前500字 ===")
print((parsed.structure_text or "")[:500])
